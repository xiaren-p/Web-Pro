import traceback
from typing import List, Optional
import json
import re
import hashlib
from datetime import timedelta
from django.utils import timezone
from django.conf import settings
import io
import os
import time
import tempfile
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAuthenticated

from api_v1.utils.responses import drf_ok, drf_error
from api_v1.utils.pagination import paginate_queryset
from api_v1.models import MonthlyLossOrder, MonthlyLossOrderFirst20
from api_v1.serializers import MonthlyLossSerializer, MonthlyLossFirst20Serializer
from django.core.cache import cache
import threading


try:
    from api_v1.models import OrderProfitCache
except Exception:
    OrderProfitCache = None

from api_v1.views.finance._helpers import (parse_months, parse_store_param, _safe_float, _safe_int, _agg_sum, _agg_int_sum, norm_month, RATIO_FIELDS, format_ratio_value, determine_month_color, build_cache_key, is_refresh_requested, try_get_cached_file_response, remove_cache_and_disk, cache_data_bytes_with_fallback, stream_tempfile_response)

class MonthlyLossViewSet(viewsets.ViewSet):
    """Monthly loss orders (CRUD + filter by month/owner, owner optional). Parameters and JSON responses use English keys only."""

    def get_permissions(self):
        method = getattr(self.request, 'method', '').upper() if hasattr(self, 'request') else ''
        if method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]

    @action(detail=False, methods=["get", "post"], url_path="")
    def list_or_create(self, request):
        if request.method.lower() == 'get':
            qs = MonthlyLossOrder.objects.all().order_by('-month', '-id')
            month_q = request.query_params.get('month')
            owner_q = request.query_params.get('owner')
            # 接受 YYYYMM 或 YYYY-MM 格式输入；匹配数据库中可能为 'YYYYMM' 或 'YYYY-MM' 的值
            def _month_variants(m):
                if not m:
                    return []
                s = str(m).strip()
                if re.match(r'^\d{6}$', s):
                    return [s, f"{s[:4]}-{s[4:6]}"]
                if re.match(r'^\d{4}-\d{2}$', s):
                    return [s, s.replace('-', '')]
                return [s]
            try:
                if month_q:
                    variants = _month_variants(month_q)
                    from django.db.models import Q
                    qf = Q()
                    for v in variants:
                        qf |= Q(month__iexact=v)
                    qs = qs.filter(qf)
            except Exception:
                pass
            try:
                if owner_q:
                    qs = qs.filter(owner__iexact=str(owner_q).strip())
            except Exception:
                pass
            total, items, _, _ = paginate_queryset(request, qs)
            data = MonthlyLossSerializer(items, many=True).data
            return drf_ok({'total': total, 'list': data})
        # POST -> 创建
        payload = request.data or {}
        ser = MonthlyLossSerializer(data=payload)
        if not ser.is_valid():
            return drf_error('参数错误', status=400, data={'errors': ser.errors})
        obj = ser.save()
        return drf_ok(MonthlyLossSerializer(obj).data, status=201)

    @action(detail=False, methods=["get"], url_path=r"(?P<id>[^/]+)/form")
    def form(self, request, id: str):
        try:
            obj = MonthlyLossOrder.objects.get(pk=id)
        except MonthlyLossOrder.DoesNotExist:
            return drf_error("未找到记录", status=404)
        return drf_ok(MonthlyLossSerializer(obj).data)

    @action(detail=False, methods=["put", "delete"], url_path=r"(?P<ids>[^/]+)")
    def update_or_delete(self, request, ids: str):
        if request.method.lower() == 'put':
            first_id = ids.split(',')[0]
            try:
                obj = MonthlyLossOrder.objects.get(pk=first_id)
            except MonthlyLossOrder.DoesNotExist:
                return drf_error("未找到记录", status=404)
            ser = MonthlyLossSerializer(obj, data=request.data or {}, partial=True)
            if not ser.is_valid():
                return drf_error('参数错误', status=400, data={'errors': ser.errors})
            obj = ser.save()
            return drf_ok(MonthlyLossSerializer(obj).data)
        # 删除
        id_list = [i for i in ids.split(',') if i]
        MonthlyLossOrder.objects.filter(id__in=id_list).delete()
        return drf_ok(status=204)

    @action(detail=False, methods=["get", "post"], url_path="download")
    def download(self, request):
        # 更简洁的单线程导出：按产品键分块遍历并写入行
        try:
            if getattr(request, 'method', '').upper() == 'GET':
                payload = request.query_params or {}
            else:
                payload = request.data or {}
            owner_q = payload.get('owner')
            time_range = payload.get('time') or payload.get('months')
            if not time_range:
                return drf_error('time parameter required, example: 202610-202612', status=400)

            months = parse_months(str(time_range))
            # 支持区间或单月输入；若 parse_months 返回空，则尝试解析单月格式
            if not months:
                single = norm_month(str(time_range))
                if single:
                    months = [single]
            if not months:
                return drf_error('please provide valid month or month range, example: 202510 or 202510-202512', status=400)

            # 准备缓存 key（保持一致），并尝试从缓存快速返回
            batch_size = int(payload.get('batch_size', 500) or 500)
            cache_key = build_cache_key(owner_q, time_range, parse_store_param(payload.get('store') or payload.get('stores')), months, batch_size)


            # 支持请求参数触发强制刷新缓存（refresh/refresh_cache/force_refresh）
            refresh_flag = is_refresh_requested(payload)
            if cache_key:
                if refresh_flag:
                    remove_cache_and_disk(cache_key)
                try:
                    if not refresh_flag:
                        resp = try_get_cached_file_response(cache_key, months)
                        if resp:
                            return resp
                except Exception:
                    pass

            # 接受 `store`（首选）或兼容旧的 `stores` 参数；
            # 过滤规则：当任意提供的店铺值包含于 `store_country` 字段时匹配
            stores = parse_store_param(payload.get('store') or payload.get('stores'))

            variant_months = set()
            for m in months:
                variant_months.add(m); variant_months.add(m.replace('-', ''))
            qs = MonthlyLossOrder.objects.filter(month__in=list(variant_months))
            if owner_q:
                qs = qs.filter(owner__iexact=str(owner_q).strip())
            if stores:
                from django.db.models import Q
                q = Q()
                for s in stores:
                    if not s:
                        continue
                    q |= Q(store_country__icontains=s)
                qs = qs.filter(q)

            try:
                if qs.count() == 0:
                    return drf_ok(status=204)
            except Exception:
                pass

            from django.db import close_old_connections

            metrics = [('sales','sales'), ('gross_profit','gross_profit'), ('gross_margin','gross_margin'), ('net_gross_margin','net_gross_margin'), ('return_rate','return_rate'), ('refund_rate','refund_amount_rate'), ('storage_fee','total_stock_fee'), ('ad_cost','spend'), ('ad_rate','spend_rate')]

            # 将 product keys 列表化，避免 MySQL 嵌套游标导致的冲突
            product_keys = list(qs.values_list('msku','asin','parent_asin','store_country').distinct())
            def chunked_iter(seq, size=500):
                batch = []
                for d in seq:
                    msku, asin, parent_asin, store_country = d
                    key = (msku or '', asin or '', parent_asin or '', store_country or '')
                    batch.append(key)
                    if len(batch) >= size:
                        yield batch
                        batch = []
                if batch:
                    yield batch


            # 单次聚合流程：一次性批量获取匹配行，在内存中聚合，然后写入 xlsx
            import tempfile, os, threading, shutil

            batch_size = int(payload.get('batch_size', 500) or 500)
            close_old_connections()

            try:
                # 一次性批量查询所有相关行（可能很大；请确保服务器有足够内存）
                rows_q = list(qs.values(
                    'msku','asin','parent_asin','store_country',
                    'product_name_sku','owner','month','sales','gross_profit','gross_margin',
                    'net_gross_margin','return_rate','refund_amount_rate','total_stock_fee','spend','spend_rate'
                ))

                # 使用 pandas 按产品与月份进行聚合（更快、更简洁）
                try:
                    import pandas as pd
                except Exception:
                    return drf_error('pandas is required for optimized aggregation, please install it', status=500)

                try:
                    df = pd.DataFrame(rows_q)
                    if df.empty:
                        aggregated = {}
                    else:
                        # 规范化字段：month -> YYYY-MM，msku/parent_asin/store_country 做清洗
                        df['month_norm'] = df.get('month').apply(norm_month)

                        # 确保数值列为数值类型
                        num_cols = ['sales','gross_profit','gross_margin','net_gross_margin','return_rate','refund_amount_rate','total_stock_fee','spend','spend_rate']
                        for c in num_cols:
                            if c in df.columns:
                                df[c] = pd.to_numeric(df[c], errors='coerce')

                        group_cols = ['msku','asin','parent_asin','store_country','month_norm']
                        agg_dict = {
                            'product_name_sku': lambda x: next((v for v in x if pd.notna(v) and str(v).strip()), ''),
                            'owner': lambda x: next((v for v in x if pd.notna(v) and str(v).strip()), ''),
                            'sales': 'sum',
                            'gross_profit': 'sum',
                            'total_stock_fee': 'sum',
                            'spend': 'sum',
                            'gross_margin': 'mean',
                            'net_gross_margin': 'mean',
                            'return_rate': 'mean',
                            'refund_amount_rate': 'mean',
                            'spend_rate': 'mean',
                        }

                        grp = df.groupby(group_cols).agg(agg_dict).reset_index()

                        # 构建与之前结构类似的聚合字典
                        aggregated = {}
                        for _, rowg in grp.iterrows():
                            try:
                                msku = rowg.get('msku') or ''
                                asin = rowg.get('asin') or ''
                                parent_asin = rowg.get('parent_asin') or ''
                                store_country = rowg.get('store_country') or ''
                                month = rowg.get('month_norm')
                                key = (msku, asin, parent_asin, store_country)
                                entry = aggregated.setdefault(key, {'meta': {'product_name': '', 'owner': ''}, 'months': {m: {} for m in months}})
                                if rowg.get('product_name_sku'):
                                    entry['meta']['product_name'] = rowg.get('product_name_sku')
                                if rowg.get('owner'):
                                    entry['meta']['owner'] = rowg.get('owner')
                                if month is None:
                                    continue
                                # 赋值聚合后的指标
                                entry['months'][month] = {
                                    'sales': int(rowg['sales']) if pd.notna(rowg['sales']) else None,
                                    'gross_profit': float(rowg['gross_profit']) if pd.notna(rowg['gross_profit']) else None,
                                    'gross_margin': float(rowg['gross_margin']) if pd.notna(rowg['gross_margin']) else None,
                                    'net_gross_margin': float(rowg['net_gross_margin']) if pd.notna(rowg['net_gross_margin']) else None,
                                    'return_rate': float(rowg['return_rate']) if pd.notna(rowg['return_rate']) else None,
                                    'refund_amount_rate': float(rowg['refund_amount_rate']) if pd.notna(rowg['refund_amount_rate']) else None,
                                    'total_stock_fee': float(rowg['total_stock_fee']) if pd.notna(rowg['total_stock_fee']) else None,
                                    'spend': float(rowg['spend']) if pd.notna(rowg['spend']) else None,
                                    'spend_rate': float(rowg['spend_rate']) if pd.notna(rowg['spend_rate']) else None,
                                }
                            except Exception:
                                continue
                except Exception:
                    # 如果 pandas 异常，则回退为空的聚合结果
                    aggregated = {}

                # 从聚合后的数据生成最终的 xlsx 文件
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                tmp_name = tmp.name
                tmp.close()

                try:
                    from openpyxl import Workbook
                except Exception:
                    try:
                        os.remove(tmp_name)
                    except Exception:
                        pass
                    return drf_error('openpyxl is required for xlsx export, please install it', status=500)

                # 使用普通 Workbook 以支持合并单元格与样式
                wb = Workbook()
                ws = wb.active
                ws.title = 'monthly_loss'

                # 中文表头映射
                basic_cols = ['MSKU','ASIN','parent_asin','store_country','product_name','owner']
                basic_map = {
                    'MSKU': 'MSKU',
                    'ASIN': 'ASIN',
                    'parent_asin': '父ASIN',
                    'store_country': '店铺',
                    'product_name': '品名/SKU',
                    'owner': '负责人',
                }
                metric_map = {
                    'sales': '销量',
                    'gross_profit': '毛利',
                    'gross_margin': '毛利率',
                    'net_gross_margin': '净毛利率',
                    'return_rate': '退货率',
                    'refund_rate': '退款率',
                    'storage_fee': '仓储费',
                    'ad_cost': '广告花费',
                    'ad_rate': '广告占比',
                }

                # 样式与边框策略：深线用于不同内容/分组分隔，浅线用于月份内列分隔，整体对比色为黑色
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                header_font = Font(bold=True)
                header_fill = PatternFill(fill_type='solid', start_color='FFDCE6F1', end_color='FFDCE6F1')
                center = Alignment(horizontal='center', vertical='center')
                # 定义边框线：深线（分组边界）、浅线（月内边界）、外框（黑色）
                dark_side = Side(border_style='medium', color='FF000000')
                light_side = Side(border_style='thin', color='FFBBBBBB')
                outer_side = Side(border_style='medium', color='FF000000')
                default_side = Side(border_style='thin', color='FF000000')

                # 预创建并缓存 Border 实例以复用，避免大量重复对象导致的比较开销
                side_map = {
                    'dark': dark_side,
                    'light': light_side,
                    'outer': outer_side,
                    'default': default_side,
                }
                _border_cache = {}
                def get_border(l_key='default', r_key='default', t_key='default', b_key='default'):
                    key = (l_key, r_key, t_key, b_key)
                    if key in _border_cache:
                        return _border_cache[key]
                    left = side_map.get(l_key, default_side)
                    right = side_map.get(r_key, default_side)
                    top = side_map.get(t_key, default_side)
                    bottom = side_map.get(b_key, default_side)
                    b = Border(left=left, right=right, top=top, bottom=bottom)
                    _border_cache[key] = b
                    return b

                # 构造两行表头并记录分组边界
                col_idx = 1
                metric_groups = []
                basic_count = len(basic_cols)
                # 先写基本列，并垂直合并两行；在基本列末尾使用深线作为分隔
                for i, c in enumerate(basic_cols):
                    ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
                    cell = ws.cell(row=1, column=col_idx, value=basic_map.get(c, c))
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center
                    # 如果是基本列最后一列，在右侧使用深线分隔；上下使用深线
                    if i == basic_count - 1:
                        cell.border = get_border('outer', 'dark', 'outer', 'outer')
                    else:
                        cell.border = get_border('outer', 'default', 'outer', 'outer')
                    col_idx += 1

                # 接着为每个指标写合并的顶层单元格，并在第二行写各月列，同时记录各组起止列
                for gi, (mname, _) in enumerate(metrics):
                    label_cn = metric_map.get(mname, mname)
                    start_col = col_idx
                    end_col = col_idx + len(months) - 1
                    metric_groups.append((start_col, end_col))
                    # 合并第一行的月份组标签
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                    top_cell = ws.cell(row=1, column=start_col, value=label_cn)
                    top_cell.font = header_font
                    top_cell.fill = header_fill
                    top_cell.alignment = center
                    # 顶层组单元格左右两端使用深线，顶部/底部使用深线
                    top_cell.border = get_border(('dark' if start_col == basic_count + 1 else 'default'), 'dark', 'outer', 'outer')
                    # 第二行写月份小标题，组内使用浅线分隔，组首使用深线左边界，组尾使用深线右边界
                    for i, m in enumerate(months):
                        cidx = start_col + i
                        cell2 = ws.cell(row=2, column=cidx, value=m)
                        cell2.font = header_font
                        cell2.fill = header_fill
                        cell2.alignment = center
                        left_key = 'dark' if i == 0 else 'light'
                        right_key = 'dark' if i == (len(months) - 1) else 'light'
                        cell2.border = get_border(left_key, right_key, 'outer', 'outer')
                    col_idx = end_col + 1

                total_cols = col_idx - 1

                # 写入数据行，从第三行开始；写入后为每个单元格应用边框样式以区分分组与月份
                rows_written = 0
                for key, entry in aggregated.items():
                    try:
                        msku, asin, parent_asin, store_country = key
                        meta = entry.get('meta', {})
                        row = [msku, asin, parent_asin, store_country, meta.get('product_name',''), meta.get('owner','')]
                        # 比率字段格式化为百分比显示
                        for _, field in metrics:
                            for m in months:
                                v = entry['months'].get(m, {}).get(field)
                                if v is None or v == '':
                                    row.append('')
                                    continue
                                if field in RATIO_FIELDS:
                                    row.append(format_ratio_value(v))
                                else:
                                    row.append(v)
                        ws.append(row)
                        rows_written += 1
                        # 应用边框：根据列位置决定左右边框是深线或浅线，顶部/底部使用默认黑色细线
                        # 避免使用 ws.max_row（在某些环境下可能触发 empty _cells 错误），使用已追踪的行号
                        r = 2 + rows_written
                        for c in range(1, total_cols + 1):
                            cell = ws.cell(row=r, column=c)
                            # 初始化左右边框为默认细黑线
                            left = default_side
                            right = default_side
                            # 若在 basic 列内部，basic 与 metrics 交界处使用深线
                            if c == basic_count:
                                border_obj = get_border('default', 'dark', 'default', 'default')
                            else:
                                # 判断是否为任一 metric group 的首列或末列
                                is_group_edge = False
                                left_key = 'default'
                                right_key = 'default'
                                for (gstart, gend) in metric_groups:
                                    if c == gstart:
                                        left_key = 'dark'
                                        is_group_edge = True
                                        break
                                    if c == gend:
                                        right_key = 'dark'
                                        is_group_edge = True
                                        break
                                if not is_group_edge:
                                    # 非边界列使用浅线分隔
                                    left_key = 'light'
                                    right_key = 'light'
                                border_obj = get_border(left_key, right_key, 'default', 'default')
                            cell.border = border_obj
                            # 对月份列（metric 列）按规则着色：对每个 month 做一次判断并为该月所有 metric 单元格上色
                            try:
                                if c > basic_count:
                                    cols_per_metric = len(months)
                                    offset = c - basic_count - 1
                                    month_idx = offset % cols_per_metric
                                    month_key = months[month_idx]
                                    month_vals = entry.get('months', {}).get(month_key, {})
                                    color = determine_month_color(month_vals)
                                    if color:
                                        try:
                                            metric_idx = offset // cols_per_metric
                                            metric_field = metrics[metric_idx][1] if 0 <= metric_idx < len(metrics) else None
                                        except Exception:
                                            metric_field = None
                                        # 不为销量、仓储费、广告花费的列才上色
                                        if metric_field in ('sales', 'total_stock_fee', 'spend'):
                                            pass
                                        else:
                                            cell.fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
                            except Exception:
                                pass
                    except Exception:
                        continue

                if rows_written == 0:
                    try:
                        total_cols = col_idx - 1
                    except Exception:
                        total_cols = len(basic_cols) + len(metrics) * len(months)
                    note = ['No data for requested months'] + [''] * (total_cols - 1)
                    try:
                        ws.append(note)
                    except Exception:
                        pass

                try:
                    wb.save(tmp_name)
                    try:
                        with open(tmp_name, 'rb') as rf:
                            data_bytes = rf.read()
                    except Exception:
                        data_bytes = None
                except Exception:
                    try:
                        os.remove(tmp_name)
                    except Exception:
                        pass
                    raise

                filename = f"monthly_loss_{months[0].replace('-','')}_{months[-1].replace('-','')}.xlsx"

                # 尝试把字节缓存到后端（使用后端默认过期策略），失败时回退到磁盘并缓存 path
                if cache_key and data_bytes is not None:
                    try:
                        cache_data_bytes_with_fallback(cache_key, data_bytes, filename)
                    except Exception:
                        pass

                # 返回临时文件流并在后台清理（封装在 helper 中）
                resp = stream_tempfile_response(tmp_name, filename)
                if resp:
                    return resp
                # 若 helper 失败，回退为原始实现
                from django.http import FileResponse
                f = open(tmp_name, 'rb')
                resp = FileResponse(f, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                resp['Content-Disposition'] = f'attachment; filename="{filename}"'
                def _cleanup_file(path_file, delay=30):
                    try:
                        time.sleep(delay)
                        try:
                            os.remove(path_file)
                        except Exception:
                            pass
                    except Exception:
                        pass
                t = threading.Thread(target=_cleanup_file, args=(tmp_name,))
                t.daemon = True
                t.start()

                return resp
            except Exception as e:
                try:
                    tmp_name
                except Exception:
                    tmp_name = None
                try:
                    if tmp_name:
                        os.remove(tmp_name)
                except Exception:
                    pass
                tb = traceback.format_exc()
                return drf_error('failed to generate xlsx', status=500, data={'msg': str(e), 'trace': tb})
        except Exception as e:
            tb = traceback.format_exc()
            return drf_error('download failed', status=500, data={'msg': str(e), 'trace': tb})


