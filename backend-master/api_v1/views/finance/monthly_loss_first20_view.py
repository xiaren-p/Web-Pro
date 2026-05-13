import traceback
from typing import List, Optional
import json
import re
import hashlib
from datetime import timedelta
from django.utils import timezone
from django.conf import settings
import io
import os
import time
import tempfile
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAuthenticated

from api_v1.utils.responses import drf_ok, drf_error
from api_v1.utils.pagination import paginate_queryset
from api_v1.models import MonthlyLossOrder, MonthlyLossOrderFirst20
from api_v1.serializers import MonthlyLossSerializer, MonthlyLossFirst20Serializer
from django.core.cache import cache
import threading


try:
    from api_v1.models import OrderProfitCache
except Exception:
    OrderProfitCache = None

from api_v1.views.finance._helpers import (parse_months, parse_store_param, _safe_float, _safe_int, _agg_sum, _agg_int_sum, norm_month, RATIO_FIELDS, format_ratio_value, determine_month_color, build_cache_key, is_refresh_requested, try_get_cached_file_response, remove_cache_and_disk, cache_data_bytes_with_fallback, stream_tempfile_response)

class MonthlyLossFirst20ViewSet(viewsets.ViewSet):
    """Monthly first-20-days loss (CRUD + filter by month/owner, owner optional). Parameters and JSON responses use English keys only."""

    def get_permissions(self):
        method = getattr(self.request, 'method', '').upper() if hasattr(self, 'request') else ''
        if method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]

    @action(detail=False, methods=["get", "post"], url_path="")
    def list_or_create(self, request):
        if request.method.lower() == 'get':
            qs = MonthlyLossOrderFirst20.objects.all().order_by('-month', '-id')
            month_q = request.query_params.get('month')
            owner_q = request.query_params.get('owner')
            # 接受 YYYYMM 或 YYYY-MM 格式输入；匹配数据库中可能为 'YYYYMM' 或 'YYYY-MM' 的值
            def _month_variants(m):
                if not m:
                    return []
                s = str(m).strip()
                if re.match(r'^\d{6}$', s):
                    return [s, f"{s[:4]}-{s[4:6]}"]
                if re.match(r'^\d{4}-\d{2}$', s):
                    return [s, s.replace('-', '')]
                return [s]
            try:
                if month_q:
                    variants = _month_variants(month_q)
                    from django.db.models import Q
                    qf = Q()
                    for v in variants:
                        qf |= Q(month__iexact=v)
                    qs = qs.filter(qf)
            except Exception:
                pass
            try:
                if owner_q:
                    qs = qs.filter(owner__iexact=str(owner_q).strip())
            except Exception:
                pass
            total, items, _, _ = paginate_queryset(request, qs)
            data = MonthlyLossFirst20Serializer(items, many=True).data
            return drf_ok({'total': total, 'list': data})
        # POST -> 创建
        payload = request.data or {}
        ser = MonthlyLossFirst20Serializer(data=payload)
        if not ser.is_valid():
            return drf_error('参数错误', status=400, data={'errors': ser.errors})
        obj = ser.save()
        return drf_ok(MonthlyLossFirst20Serializer(obj).data, status=201)

    @action(detail=False, methods=["get"], url_path=r"(?P<id>[^/]+)/form")
    def form(self, request, id: str):
        try:
            obj = MonthlyLossOrderFirst20.objects.get(pk=id)
        except MonthlyLossOrderFirst20.DoesNotExist:
            return drf_error("未找到记录", status=404)
        return drf_ok(MonthlyLossFirst20Serializer(obj).data)

    @action(detail=False, methods=["put", "delete"], url_path=r"(?P<ids>[^/]+)")
    def update_or_delete(self, request, ids: str):
        if request.method.lower() == 'put':
            first_id = ids.split(',')[0]
            try:
                obj = MonthlyLossOrderFirst20.objects.get(pk=first_id)
            except MonthlyLossOrderFirst20.DoesNotExist:
                return drf_error("未找到记录", status=404)
            ser = MonthlyLossFirst20Serializer(obj, data=request.data or {}, partial=True)
            if not ser.is_valid():
                return drf_error('参数错误', status=400, data={'errors': ser.errors})
            obj = ser.save()
            return drf_ok(MonthlyLossFirst20Serializer(obj).data)
        # 删除
        id_list = [i for i in ids.split(',') if i]
        MonthlyLossOrderFirst20.objects.filter(id__in=id_list).delete()
        return drf_ok(status=204)

    @action(detail=False, methods=["post"], url_path="download")
    def download(self, request):
        # 导出对比：本月前20天数据 vs 上个月整月数据
        try:
            if getattr(request, 'method', '').upper() == 'GET':
                payload = request.query_params or {}
            else:
                payload = request.data or {}

            owner_q = payload.get('owner')
            time_range = payload.get('time') or payload.get('months')
            if not time_range:
                return drf_error('time parameter required, example: 202510', status=400)

            months = parse_months(str(time_range))
            # 支持单月输入（YYYYMM 或 YYYY-MM）以及区间输入
            if not months:
                single = norm_month(str(time_range))
                if single:
                    months = [single]
            # 仅支持单月对比
            if not months or len(months) != 1:
                return drf_error('please provide a single month, format: 202510 or 2025-10', status=400)
            cur_month = months[0]
            # 计算上一个月
            try:
                y = int(cur_month[:4]); m = int(cur_month[-2:])
                pm_y = y if m > 1 else y - 1
                pm_m = m - 1 if m > 1 else 12
                prev_month = f"{pm_y:04d}-{pm_m:02d}"
            except Exception:
                return drf_error('failed to compute previous month', status=400)

            # 准备 cache key（以当前月为主）
            batch_size = int(payload.get('batch_size', 500) or 500)
            cache_key = build_cache_key(owner_q, time_range, parse_store_param(payload.get('store') or payload.get('stores')), [cur_month], batch_size)

            refresh_flag = is_refresh_requested(payload)
            if cache_key and refresh_flag:
                remove_cache_and_disk(cache_key)
            if cache_key and not refresh_flag:
                try:
                    resp = try_get_cached_file_response(cache_key, [cur_month])
                    if resp:
                        return resp
                except Exception:
                    pass

            # 过滤条件
            stores = parse_store_param(payload.get('store') or payload.get('stores'))
            def make_variant_set(m):
                return {m, m.replace('-', '')}

            qs_cur = MonthlyLossOrderFirst20.objects.filter(month__in=list(make_variant_set(cur_month)))
            qs_prev = MonthlyLossOrder.objects.filter(month__in=list(make_variant_set(prev_month)))
            if owner_q:
                qs_cur = qs_cur.filter(owner__iexact=str(owner_q).strip())
                qs_prev = qs_prev.filter(owner__iexact=str(owner_q).strip())
            if stores:
                from django.db.models import Q
                qf = Q()
                for s in stores:
                    if not s:
                        continue
                    qf |= Q(store_country__icontains=s)
                qs_cur = qs_cur.filter(qf)
                qs_prev = qs_prev.filter(qf)

            # 查询并聚合（使用 pandas 与之前相同策略）
            try:
                import pandas as pd
            except Exception:
                return drf_error('pandas is required for optimized aggregation, please install it', status=500)

            rows_cur = list(qs_cur.values('msku','asin','parent_asin','store_country','product_name_sku','owner','month','sales','gross_profit','gross_margin','net_gross_margin','return_rate','refund_amount_rate','total_stock_fee','spend','spend_rate'))
            rows_prev = list(qs_prev.values('msku','asin','parent_asin','store_country','product_name_sku','owner','month','sales','gross_profit','gross_margin','net_gross_margin','return_rate','refund_amount_rate','total_stock_fee','spend','spend_rate'))

            def aggregate_rows(rows, months_list):
                try:
                    df = pd.DataFrame(rows)
                    if df.empty:
                        return {}
                    df['month_norm'] = df.get('month').apply(norm_month)
                    num_cols = ['sales','gross_profit','gross_margin','net_gross_margin','return_rate','refund_amount_rate','total_stock_fee','spend','spend_rate']
                    for c in num_cols:
                        if c in df.columns:
                            df[c] = pd.to_numeric(df[c], errors='coerce')
                    group_cols = ['msku','asin','parent_asin','store_country','month_norm']
                    agg_dict = {
                        'product_name_sku': lambda x: next((v for v in x if pd.notna(v) and str(v).strip()), ''),
                        'owner': lambda x: next((v for v in x if pd.notna(v) and str(v).strip()), ''),
                        'sales': 'sum',
                        'gross_profit': 'sum',
                        'total_stock_fee': 'sum',
                        'spend': 'sum',
                        'gross_margin': 'mean',
                        'net_gross_margin': 'mean',
                        'return_rate': 'mean',
                        'refund_amount_rate': 'mean',
                        'spend_rate': 'mean',
                    }
                    grp = df.groupby(group_cols).agg(agg_dict).reset_index()
                    out = {}
                    for _, rowg in grp.iterrows():
                        try:
                            msku = rowg.get('msku') or ''
                            asin = rowg.get('asin') or ''
                            parent_asin = rowg.get('parent_asin') or ''
                            store_country = rowg.get('store_country') or ''
                            month = rowg.get('month_norm')
                            key = (msku, asin, parent_asin, store_country)
                            entry = out.setdefault(key, {'meta': {'product_name': '', 'owner': ''}, 'months': {}})
                            if rowg.get('product_name_sku'):
                                entry['meta']['product_name'] = rowg.get('product_name_sku')
                            if rowg.get('owner'):
                                entry['meta']['owner'] = rowg.get('owner')
                            if month is None:
                                continue
                            entry['months'][month] = {
                                'sales': int(rowg['sales']) if pd.notna(rowg['sales']) else None,
                                'gross_profit': float(rowg['gross_profit']) if pd.notna(rowg['gross_profit']) else None,
                                'gross_margin': float(rowg['gross_margin']) if pd.notna(rowg['gross_margin']) else None,
                                'net_gross_margin': float(rowg['net_gross_margin']) if pd.notna(rowg['net_gross_margin']) else None,
                                'return_rate': float(rowg['return_rate']) if pd.notna(rowg['return_rate']) else None,
                                'refund_amount_rate': float(rowg['refund_amount_rate']) if pd.notna(rowg['refund_amount_rate']) else None,
                                'total_stock_fee': float(rowg['total_stock_fee']) if pd.notna(rowg['total_stock_fee']) else None,
                                'spend': float(rowg['spend']) if pd.notna(rowg['spend']) else None,
                                'spend_rate': float(rowg['spend_rate']) if pd.notna(rowg['spend_rate']) else None,
                            }
                        except Exception:
                            continue
                    return out
                except Exception:
                    return {}

            agg_cur = aggregate_rows(rows_cur, [cur_month])
            agg_prev = aggregate_rows(rows_prev, [prev_month])

            # 生成 xlsx：基本列 + 每个指标 Prev/Cur
            try:
                from openpyxl import Workbook
            except Exception:
                return drf_error('openpyxl is required for xlsx export, please install it', status=500)

            wb = Workbook()
            ws = wb.active
            ws.title = 'monthly_loss_first20_compare'

            basic_cols = ['MSKU','ASIN','parent_asin','store_country','product_name','owner']
            basic_map = {'MSKU':'MSKU','ASIN':'ASIN','parent_asin':'父ASIN','store_country':'店铺','product_name':'品名/SKU','owner':'负责人'}
            metric_map = {
                'sales': '销量',
                'gross_profit': '毛利',
                'gross_margin': '毛利率',
                'net_gross_margin': '净毛利率',
                'return_rate': '退货率',
                'refund_rate': '退款率',
                'storage_fee': '仓储费',
                'ad_cost': '广告花费',
                'ad_rate': '广告占比',
            }
            metrics = [('sales','sales'),('gross_profit','gross_profit'),('gross_margin','gross_margin'),('net_gross_margin','net_gross_margin'),('return_rate','return_rate'),('refund_rate','refund_amount_rate'),('storage_fee','total_stock_fee'),('ad_cost','spend'),('ad_rate','spend_rate')]

            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_font = Font(bold=True)
            header_fill = PatternFill(fill_type='solid', start_color='FFDCE6F1', end_color='FFDCE6F1')
            center = Alignment(horizontal='center', vertical='center')
            default_side = Side(border_style='thin', color='FF000000')
            light_side = Side(border_style='thin', color='FFBBBBBB')
            dark_side = Side(border_style='medium', color='FF000000')
            side_map = {'dark': dark_side, 'light': light_side, 'default': default_side}
            _border_cache = {}
            def get_border(l='default', r='default', t='default', b='default'):
                key = (l,r,t,b)
                if key in _border_cache:
                    return _border_cache[key]
                bobj = Border(left=side_map.get(l, default_side), right=side_map.get(r, default_side), top=side_map.get(t, default_side), bottom=side_map.get(b, default_side))
                _border_cache[key]=bobj
                return bobj

            # 写表头：两行，第一行为指标名（每指标合并两列），第二行为 Prev/Cur 小标题
            col = 1
            basic_count = len(basic_cols)
            for i,c in enumerate(basic_cols):
                ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
                cell = ws.cell(row=1, column=col, value=basic_map.get(c,c))
                cell.font = header_font; cell.fill = header_fill; cell.alignment = center
                cell.border = get_border('outer','outer','outer','outer') if i==basic_count-1 else get_border('default','default','default','default')
                col += 1

            for mname, _ in metrics:
                start = col
                end = col + 2 - 1
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
                label_cn = metric_map.get(mname, mname)
                top = ws.cell(row=1, column=start, value=label_cn)
                top.font = header_font; top.fill = header_fill; top.alignment = center
                # second row: Prev, Cur (不再输出 Diff)
                labels = [prev_month, cur_month]
                for i, lab in enumerate(labels):
                    cidx = start + i
                    cell2 = ws.cell(row=2, column=cidx, value=lab)
                    cell2.font = header_font; cell2.fill = header_fill; cell2.alignment = center
                    cell2.border = get_border('default','default','default','default')
                col = end + 1

            total_cols = col -1

            # 合并键集合
            keys = set(list(agg_cur.keys()) + list(agg_prev.keys()))
            rows_written = 0
            for key in keys:
                try:
                    msku, asin, parent_asin, store_country = key
                    meta = (agg_cur.get(key, {}) .get('meta') or agg_prev.get(key, {}).get('meta') or {})
                    row = [msku, asin, parent_asin, store_country, meta.get('product_name',''), meta.get('owner','')]
                    for _, field in metrics:
                        prev_v = agg_prev.get(key, {}).get('months', {}).get(prev_month, {}).get(field)
                        cur_v = agg_cur.get(key, {}).get('months', {}).get(cur_month, {}).get(field)
                        if field in RATIO_FIELDS:
                            pv = format_ratio_value(prev_v) if prev_v is not None else ''
                            cv = format_ratio_value(cur_v) if cur_v is not None else ''
                            row.extend([pv, cv])
                        else:
                            pv = '' if prev_v is None else prev_v
                            cv = '' if cur_v is None else cur_v
                            row.extend([pv, cv])
                    ws.append(row)
                    rows_written += 1
                    r = 2 + rows_written
                    for c in range(1, total_cols + 1):
                        cell = ws.cell(row=r, column=c)
                        # 计算左右边框：basic 列与 metrics 分组交接处使用深线，组内使用浅线
                        try:
                            if c <= basic_count:
                                left_key = 'default'
                                right_key = 'dark' if c == basic_count else 'default'
                            else:
                                # metric 区域，每个 metric 占两列 (Prev, Cur)
                                offset = c - basic_count - 1
                                metric_idx = offset // 2
                                pos_in_metric = offset % 2  # 0:Prev,1:Cur
                                left_key = 'dark' if pos_in_metric == 0 else 'light'
                                right_key = 'dark' if pos_in_metric == 1 else 'light'
                            cell.border = get_border(left_key, right_key, 'default', 'default')
                        except Exception:
                            try:
                                cell.border = get_border('default','default','default','default')
                            except Exception:
                                pass
                        # 对 Prev/Cur 单元格按规则着色（排除部分字段）
                        try:
                            if c > basic_count:
                                offset = c - basic_count - 1
                                metric_idx = offset // 2
                                col_pos = offset % 2  # 0:Prev,1:Cur
                                field_name = metrics[metric_idx][1] if 0 <= metric_idx < len(metrics) else None
                                if field_name and field_name not in ('sales', 'total_stock_fee', 'spend') and col_pos in (0,1):
                                    month_vals = None
                                    try:
                                        if col_pos == 0:
                                            month_vals = agg_prev.get(key, {}).get('months', {}).get(prev_month, {})
                                        else:
                                            month_vals = agg_cur.get(key, {}).get('months', {}).get(cur_month, {})
                                    except Exception:
                                        month_vals = None
                                    if month_vals:
                                        color = determine_month_color(month_vals)
                                        if color:
                                            try:
                                                cell.fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
                                            except Exception:
                                                pass
                        except Exception:
                            pass
                except Exception:
                    continue

            if rows_written == 0:
                note = ['No data for requested month'] + [''] * (total_cols - 1)
                ws.append(note)

            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            tmp_name = tmp.name; tmp.close()
            try:
                wb.save(tmp_name)
                with open(tmp_name, 'rb') as rf:
                    data_bytes = rf.read()
            except Exception:
                try:
                    os.remove(tmp_name)
                except Exception:
                    pass
                return drf_error('failed to generate xlsx', status=500)

            filename = f"monthly_loss_first20_compare_{cur_month.replace('-','')}.xlsx"
            if cache_key and data_bytes is not None:
                try:
                    cache_data_bytes_with_fallback(cache_key, data_bytes, filename)
                except Exception:
                    pass

            resp = stream_tempfile_response(tmp_name, filename)
            if resp:
                return resp
            from django.http import FileResponse
            f = open(tmp_name, 'rb')
            resp = FileResponse(f, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="{filename}"'
            def _cleanup(path_file, delay=30):
                try:
                    time.sleep(delay)
                    try:
                        os.remove(path_file)
                    except Exception:
                        pass
                except Exception:
                    pass
            t = threading.Thread(target=_cleanup, args=(tmp_name,))
            t.daemon = True; t.start()
            return resp
        except Exception as e:
            tb = traceback.format_exc()
            return drf_error('download failed', status=500, data={'msg': str(e), 'trace': tb})


__all__ = ["StatisticsViewSet", "MonthlyLossViewSet", "MonthlyLossFirst20ViewSet"]

