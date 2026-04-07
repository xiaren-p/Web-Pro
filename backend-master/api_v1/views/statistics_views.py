import traceback
from typing import List, Optional
import json
import re
import hashlib
from datetime import timedelta
from django.utils import timezone
from django.conf import settings
import io
import os
import time
import tempfile
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import AllowAny, IsAuthenticated

from api_v1.utils.responses import drf_ok, drf_error
from api_v1.utils.pagination import paginate_queryset
from api_v1.models import MonthlyLossOrder, MonthlyLossOrderFirst20
from api_v1.serializers import MonthlyLossSerializer, MonthlyLossFirst20Serializer
from django.core.cache import cache
import threading


try:
    from openapi.client import LingXingClient
except Exception:
    LingXingClient = None
try:
    from api_v1.models import OrderProfitCache
except Exception:
    OrderProfitCache = None


# 辅助工具移动到模块级以保持 `download` 函数简洁
def parse_months(r: str):
    try:
        if isinstance(r, str) and '-' in r:
            a, b = r.split('-', 1)
            def to_ym(x):
                y = int(x[:4]); m = int(x[4:6]); return y, m
            ys, ms = to_ym(a); ye, me = to_ym(b)
            months = []
            y, m = ys, ms
            while (y < ye) or (y == ye and m <= me):
                months.append(f"{y:04d}-{m:02d}")
                m += 1
                if m > 12:
                    m = 1; y += 1
            return months
        return []
    except Exception:
        return []


def parse_store_param(store_param):
    stores = []
    if not store_param:
        return stores
    try:
        if isinstance(store_param, str):
            try:
                parsed = json.loads(store_param)
                if isinstance(parsed, list):
                    stores = [str(x).strip() for x in parsed if x is not None and str(x).strip()]
                else:
                    stores = [s.strip() for s in store_param.split(',') if s.strip()]
            except Exception:
                stores = [s.strip() for s in store_param.split(',') if s.strip()]
        elif isinstance(store_param, (list, tuple, set)):
            stores = [str(x).strip() for x in store_param if x is not None and str(x).strip()]
        else:
            stores = [str(store_param).strip()]
    except Exception:
        stores = []
    return stores


def _safe_float(x):
    try:
        if x is None or (isinstance(x, str) and str(x).strip() == ''):
            return None
        return float(x)
    except Exception:
        try:
            return float(str(x).strip())
        except Exception:
            return None


def _safe_int(x):
    try:
        if x is None or (isinstance(x, str) and str(x).strip() == ''):
            return None
        return int(x)
    except Exception:
        try:
            return int(float(x))
        except Exception:
            return None


def _agg_sum(a, b):
    if a is None and b is None:
        return None
    if a is None:
        return b
    if b is None:
        return a
    return a + b


def _agg_int_sum(a, b):
    s = _agg_sum(a, b)
    if s is None:
        return None
    try:
        return int(s)
    except Exception:
        try:
            return int(float(s))
        except Exception:
            return None


def norm_month(x):
    try:
        if x is None:
            return None
        s = str(x).strip()
        if re.match(r'^\d{6}$', s):
            return f"{s[:4]}-{s[4:6]}"
        return s
    except Exception:
        return None


RATIO_FIELDS = {'gross_margin', 'net_gross_margin', 'return_rate', 'refund_amount_rate', 'spend_rate'}


def format_ratio_value(v):
    if v is None or v == '':
        return ''
    try:
        fv = float(v)
        if abs(fv) <= 1:
            fv = fv * 100.0
        return f"{fv:.2f}%"
    except Exception:
        return str(v)


def determine_month_color(month_vals):
    try:
        gp = month_vals.get('gross_profit')
        refund = month_vals.get('refund_amount_rate')
        ad_rate = month_vals.get('spend_rate')

        def _norm_pct(v):
            try:
                if v is None:
                    return None
                fv = float(v)
                if abs(fv) <= 1:
                    return fv * 100.0
                return fv
            except Exception:
                return None

        gp_v = None
        try:
            gp_v = float(gp) if gp is not None else None
        except Exception:
            gp_v = None
        refund_v = _norm_pct(refund)
        ad_v = _norm_pct(ad_rate)
        # 规则判断：均要求毛利润 < 0
        if gp_v is not None and gp_v < 0:
            if refund_v is not None and ad_v is not None and refund_v > 15 and ad_v > 10:
                return 'FFFF0000'  # red
            if refund_v is not None and refund_v > 15 and (ad_v is None or ad_v <= 10):
                return 'FF00AA00'  # green
            if refund_v is not None and refund_v <= 15 and ad_v is not None and ad_v > 10:
                return 'FFFFFF00'  # yellow
        return None
    except Exception:
        return None


# 模块级辅助函数：缓存与响应处理（从 download 中抽取，便于复用和单元测试）
def build_cache_key(owner_q, time_range, stores, months, batch_size):
    try:
        cache_input = {
            'owner': owner_q,
            'time': time_range,
            'stores': stores,
            'months': months,
            'batch_size': batch_size,
        }
        cache_raw = json.dumps(cache_input, ensure_ascii=False, sort_keys=True)
        return 'monthly_loss_xlsx_' + hashlib.sha256(cache_raw.encode('utf-8')).hexdigest()
    except Exception:
        return None


def is_refresh_requested(payload) -> bool:
    try:
        for rf_key in ('refresh', 'refresh_cache', 'force_refresh'):
            v = payload.get(rf_key)
            if isinstance(v, str):
                if v.strip().lower() in ('1', 'true', 'yes', 'y', 'on'):
                    return True
            elif v:
                return True
    except Exception:
        pass
    return False


def try_get_cached_file_response(cache_key, months):
    try:
        if not cache_key:
            return None
        cached = cache.get(cache_key)
        if not cached:
            return None
        # dict with 'data'
        if isinstance(cached, dict) and cached.get('data'):
            data_bytes = cached.get('data')
            filename = cached.get('filename') or f"monthly_loss_{months[0].replace('-','')}_{months[-1].replace('-','')}.xlsx"
            bio = io.BytesIO(data_bytes)
            from django.http import FileResponse
            resp = FileResponse(bio, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="{filename}"'
            return resp
        # dict with 'path'
        if isinstance(cached, dict) and cached.get('path'):
            p = cached.get('path')
            try:
                if p and os.path.exists(p):
                    f = open(p, 'rb')
                    from django.http import FileResponse
                    resp = FileResponse(f, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    resp['Content-Disposition'] = f'attachment; filename="{cached.get("filename") or os.path.basename(p)}"'
                    return resp
            except Exception:
                return None
        # raw bytes
        if isinstance(cached, (bytes, bytearray)):
            bio = io.BytesIO(bytes(cached))
            from django.http import FileResponse
            resp = FileResponse(bio, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="monthly_loss_{months[0].replace('-','')}_{months[-1].replace('-','')}.xlsx"'
            return resp
    except Exception:
        return None
    return None


def remove_cache_and_disk(cache_key):
    try:
        if not cache_key:
            return
        old = None
        try:
            old = cache.get(cache_key)
        except Exception:
            old = None
        try:
            cache.delete(cache_key)
        except Exception:
            pass
        try:
            if isinstance(old, dict) and old.get('path'):
                pth = old.get('path')
                try:
                    if pth and os.path.exists(pth):
                        os.remove(pth)
                except Exception:
                    pass
        except Exception:
            pass
        try:
            media_root = getattr(settings, 'MEDIA_ROOT', None) or os.path.join(os.path.dirname(__file__), '..', 'media')
            export_dir = os.path.join(media_root, 'export_cache')
            file_path = os.path.join(export_dir, f"{cache_key}.xlsx")
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception:
                pass
        except Exception:
            pass
    except Exception:
        pass


def cache_data_bytes_with_fallback(cache_key, data_bytes, filename):
    try:
        if not cache_key or data_bytes is None:
            return False
        try:
            ok = cache.set(cache_key, {'data': data_bytes, 'filename': filename}, timeout=None)
        except TypeError:
            ok = cache.set(cache_key, {'data': data_bytes, 'filename': filename})
        # 某些缓存后端（Django core/backends）返回 None 表示成功，仅当明确返回 False 时视为失败
        if ok is not False:
            return True
        # 回退到磁盘
        try:
            media_root = getattr(settings, 'MEDIA_ROOT', None) or os.path.join(os.path.dirname(__file__), '..', 'media')
            export_dir = os.path.join(media_root, 'export_cache')
            os.makedirs(export_dir, exist_ok=True)
            file_path = os.path.join(export_dir, f"{cache_key}.xlsx")
            with open(file_path, 'wb') as wf:
                wf.write(data_bytes)
            try:
                cache.set(cache_key, {'path': file_path, 'filename': filename}, timeout=None)
            except TypeError:
                cache.set(cache_key, {'path': file_path, 'filename': filename})
            return True
        except Exception:
            # 写磁盘回退失败则继续由外层捕获并返回 False
            pass
        return False
    except Exception:
        return False


def stream_tempfile_response(tmp_name, filename, delay=30):
    try:
        from django.http import FileResponse
        f = open(tmp_name, 'rb')
        resp = FileResponse(f, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="{filename}"'

        def _cleanup_file(path_file, delay_sec=delay):
            try:
                time.sleep(delay_sec)
                try:
                    os.remove(path_file)
                except Exception:
                    pass
            except Exception:
                pass

        t = threading.Thread(target=_cleanup_file, args=(tmp_name,))
        t.daemon = True
        t.start()
        return resp
    except Exception:
        return None

class StatisticsViewSet(viewsets.ViewSet):
    """统计类接口（包括 lossmaking orders 查询）"""

    def get_permissions(self):
        method = getattr(self.request, 'method', '').upper() if hasattr(self, 'request') else ''
        if method in ('GET', 'POST'):
            return [AllowAny()]
        return [IsAuthenticated()]

    # 模块级辅助函数：供多个 endpoint 和后台刷新的线程复用
    @staticmethod
    def parse_number(v) -> Optional[float]:
        """
        将各种可能的数值表示安全地解析为 float。

        支持输入类型：int、float、带千分位逗号的字符串（例如 "1,234.56"）、空字符串或 None。
        解析规则：
        - 若输入为 None 或空字符串，返回 None；
        - 若输入为 int/float，返回对应的 float 值；
        - 若输入为字符串，则移除逗号并尝试转换为 float，转换失败返回 None。

        该函数用于解析上游接口返回的可能为字符串的数值字段（例如 `gross_margin`），
        并在遇到异常数据时避免抛出异常影响整体流程。

        :param v: 待解析的值
        :return: 解析后的 float，或解析失败时的 None
        """
        try:
            if v is None:
                return None
            if isinstance(v, (int, float)):
                return float(v)
            s = str(v).replace(',', '').strip()
            if s == '':
                return None
            return float(s)
        except Exception:
            return None

    @staticmethod
    def pick_fields(item: dict) -> dict:
        """
        从上游单条记录中抽取并规范化需要缓存和返回给前端的字段集合。

        目的：减少缓存体积，隔离上游字段变动对前端的影响，同时只保留前端关心的字段。

        返回字段示例：`gross_margin`, `gross_profit`, `currency_icon`, `sids`, `price_list`,
        `seller_store_countries`（只保留国家名），以及局部的 `local_infos`（只保留 name/sku）。

        :param item: 上游返回的单条记录
        :return: 仅包含必要字段的字典
        """
        return {
            'gross_margin': item.get('gross_margin'),
            'gross_profit': item.get('gross_profit'),
            'currency_icon': item.get('currency_icon'),
            'principal_names': item.get('principal_names'),
            'sids': item.get('sids'),
            'small_image_url': item.get('small_image_url'),
            'currency_code': item.get('currency_code'),
            'price_list': item.get('price_list') or [],
            'parent_asins': item.get('parent_asins') or [],
            'seller_store_countries': [c.get('name') for c in (item.get('seller_store_countries') or []) if c.get('name')],
            'local_infos': [
                {'local_name': li.get('local_name'), 'local_sku': li.get('local_sku')} for li in (item.get('local_infos') or [])
            ],
            'net_gross_margin': item.get('net_gross_margin'),
            'return_rate': item.get('return_rate'),
            'refund_amount_rate': item.get('refund_amount_rate'),
            'total_stock_fee': item.get('total_stock_fee'),
            'spend': item.get('spend'),
            'spend_rate': item.get('spend_rate'),
        }


    def apply_rule(self, items: List[dict], rule_name: Optional[str]) -> List[dict]:
        """
        根据传入的 `rule_name` 对 items 做规则化处理或过滤。

        说明：当前方法为占位实现（no-op），保留多个 rule 分支以便未来扩展具体规则逻辑。
        如果没有提供 `rule_name` 或遇到异常，会直接返回原始 items。

        :param items: 待处理的条目列表（通常为 pick_fields 之后的条目）
        :param rule_name: 规则名称（字符串），例如 'rule1' 等
        :return: 经规则处理后的条目列表
        """
        if not rule_name:
            return items
        out: List[dict] = []
        for it in items:
            try:
                gm = self.parse_number(it.get('gross_profit'))
                refund = self.parse_number(it.get('refund_amount_rate'))
                spend = self.parse_number(it.get('spend_rate'))
                cond1 = (gm is not None and gm < 0)
                cond2 = (refund is not None and refund > 0.15)
                cond3 = (spend is not None and spend > 0.10)
                match = False
                if rule_name == 'rule1':
                    match = cond1 or cond2 or cond3
                elif rule_name == 'rule2':
                    match = cond1 and cond2 and cond3
                elif rule_name == 'rule3':
                    match = cond1 and cond2 and not cond3
                elif rule_name == 'rule4':
                    match = cond1 and cond3 and not cond2
                if match:
                    out.append(it)
            except Exception:
                # 单条解析异常跳过
                continue
        return out

    def _run_background_refresh(self, cache_key, key_obj, payload, lock_key):
        """
        后台刷新任务：按 sids 分片、分页向上游请求数据、过滤并写入本地缓存。
        行为要点：
            - 按 `payload['sids']` 将 sids 列表切分为若干 chunk（每 chunk 最多 50 个 sids），
              若未提供 sids，则向上游请求全量（通过 None 表示不传 sids）；
            - 对每个 chunk，使用固定的 `length`（分页大小）从上游拉取所有页的数据；
            - 对每一页数据应用 `filter_gross_margin`，只保留亏损条目；汇总所有 chunk 的结果为 `bg_all`；
            - 将 `bg_all` 中的每项通过 `pick_fields` 规范化后写入 `OrderProfitCache`（以 `cache_key` 为键）；
            - 无论成功或失败，最后都会删除分布式锁 `lock_key`，以释放刷新权限。
            参数说明：
            - cache_key: 用于在本地 DB 缓存中定位记录的哈希键；
            - key_obj: 用于记录 params 的原始对象（会序列化保存到缓存的 `params` 字段）；
            - payload: 原始请求负载（包含 startDate/endDate/currencyCode/sids/rule 等）；
            - lock_key: 分布式锁的缓存 key，函数结束时必须删除该锁。
        注意：该函数在后台线程中执行，必须对所有异常进行防护，避免抛出未捕获异常。
        """
        try:
            client = LingXingClient() if LingXingClient is not None else None
            length = 5000
            bg_all = []
            bg_last = None

            # 不再按 sids 分片：后台刷新对上游请求不携带 sids（全量请求）
            sids_chunks_local = [None]

            def do_request_once_local(client_local, offset_val, chunk_param):
                """
                向上游请求单页数据的内部函数。
                - 如果 `chunk_param` 有值，则在请求体中添加 `sids` 字段；否则不传 `sids`，表示请求全量数据（或由上游决定默认行为）。
                - 返回值为 (page_filtered, info)，其中 page_filtered 是经过 `filter_gross_margin` 筛选的本页数据，
                  info 包含本次调用的统计信息（code/message/this_data_len/total）。
                """
                body = {
                    'offset': offset_val,
                    'length': length,
                    'startDate': payload.get('startDate'),
                    'endDate': payload.get('endDate'),
                }
                if payload.get('currencyCode'):
                    body['currencyCode'] = payload.get('currencyCode')
                resp = None
                try:
                    # 不设置超时，由后台线程控制请求时长（取消上游请求的超时限制）
                    resp = client_local.request_sync('/basicOpen/finance/mreport/OrderProfit', 'POST', req_body=body)
                except Exception as e:
                    tb = traceback.format_exc()
                    raise
                this_data = getattr(resp, 'data', []) or []
                try:
                    info = {
                        'code': int(getattr(resp, 'code', -1) or -1),
                        'message': getattr(resp, 'message', None),
                        'this_data_len': len(this_data),
                        'total': getattr(resp, 'total', None),
                    }
                except Exception:
                    # 记录 resp 的简要表示，便于排查 info 构造失败的情况
                    try:
                        resp_repr = repr(resp)
                    except Exception:
                        resp_repr = str(resp)[:1000]
                    info = {'raw': resp_repr[:1000]}
                return this_data, info

            import math
            for idx, chunk in enumerate(sids_chunks_local):
                try:
                    first_page_data, first_info = do_request_once_local(client, 0, chunk)
                except Exception:
                    tb = traceback.format_exc()
                    continue
                if first_page_data:
                    bg_all.extend(first_page_data)
                if first_info:
                    bg_last = first_info

                try:
                    total_available = int(first_info.get('total')) if first_info and first_info.get('total') is not None else None
                except Exception:
                    total_available = None

                if not total_available or total_available <= len(first_page_data):
                    continue

                total_pages = math.ceil(total_available / length)
                remaining_offsets = [i * length for i in range(1, total_pages)]
                for off in remaining_offsets:
                    try:
                        data, info = do_request_once_local(client, off, chunk)
                        if data:
                            bg_all.extend(data)
                        if info:
                            bg_last = info
                    except Exception as e:
                        tb = traceback.format_exc()
                        continue

            # 构造 pick_fields 并写入 OrderProfitCache
            try:
                picked_bg = [self.pick_fields(it) for it in bg_all] if bg_all else []
                if cache_key and OrderProfitCache is not None and picked_bg:
                    try:
                        obj, created = OrderProfitCache.objects.update_or_create(
                            key=cache_key,
                            defaults={
                                'params': json.dumps(key_obj, ensure_ascii=False),
                                'data': json.dumps(picked_bg, ensure_ascii=False),
                            }
                        )
                        try:
                            _ = obj.created_at.isoformat()
                        except Exception:
                            pass
                    except Exception:
                        pass
            except Exception:
                pass
        finally:
            try:
                cache.delete(lock_key)
            except Exception:
                pass

    @action(detail=False, methods=["post"], url_path="lossmakingorders_sync")
    def lossmaking_orders_sync(self, request):
        """返回本次查询应使用的缓存 key 与上次同步时间，并告知是否需要刷新缓存。

        请求 body (JSON): 同 `lossmakingorders`（sids/startDate/endDate/currencyCode/rule）
        返回: { key: str, sync_time: str|null, needs_refresh: bool }
        """
        # 说明：该接口只负责判断与触发后台刷新（若需要），不直接返回业务数据。
        # 返回字段解释：
        # - key: 本次查询对应的缓存 key，供前端后续调用 `lossmakingorders_data` 使用；
        # - sync_time: 上次缓存写入时间（ISO 格式），若无缓存则为 null；
        # - needs_refresh: 若缓存不存在或超过 10 分钟则为 True；若最近已刷新则为 False；
        # - syncing: 如果本次触发了后台刷新或已有其他进程在刷新，则为 True（表示正在同步）。
        try:
            if getattr(request, 'method', '').upper() == 'GET':
                payload = request.query_params or {}
            else:
                payload = request.data or {}

            # sids 可选：sync 阶段不再依赖前端传入的 sids，缓存与上游请求均不携带 sids
            sids = payload.get('sids')
            if isinstance(sids, str):
                try:
                    sids = json.loads(sids)
                except Exception:
                    try:
                        sids = [int(x) for x in sids.split(',') if x.strip()]
                    except Exception:
                        sids = [x.strip() for x in sids.split(',') if x.strip()]

            start_date = payload.get('startDate')
            end_date = payload.get('endDate')
            if not (start_date and end_date):
                return drf_error('startDate and endDate are required', status=400)

            currency_code = payload.get('currencyCode')

            try:
                key_obj = {
                    'startDate': start_date,
                    'endDate': end_date,
                    'currencyCode': currency_code,
                }
                key_raw = json.dumps(key_obj, ensure_ascii=False, sort_keys=True)
                cache_key = hashlib.sha256(key_raw.encode('utf-8')).hexdigest()
            except Exception:
                return drf_error('failed to build cache key', status=500)

            cache_created_at = None
            needs_refresh = True
            try:
                if cache_key and OrderProfitCache is not None:
                    entry = OrderProfitCache.objects.filter(key=cache_key).first()
                    if entry:
                        try:
                            # 使用 updated_at 判断缓存是否过期（update_or_create 会刷新 updated_at）
                            cache_created_at = entry.updated_at.isoformat()
                            if timezone.now() - entry.updated_at <= timedelta(minutes=10):
                                needs_refresh = False
                        except Exception:
                            pass
            except Exception:
                pass

            # 若需要刷新（缓存缺失或过期），尝试获取锁并在后台启动刷新任务
            syncing_flag = False
            try:
                if needs_refresh and cache_key:
                    lock_key = f'OrderProfit_lock_{cache_key}'
                    got_lock = False
                    try:
                        got_lock = cache.add(lock_key, '1', timeout=600)
                    except Exception:
                        got_lock = False

                    if got_lock:
                        syncing_flag = True
                        try:
                            # 使用独立守护线程执行后台刷新，避免临时 ThreadPoolExecutor 被回收导致任务被取消
                            t = threading.Thread(target=self._run_background_refresh, args=(cache_key, key_obj, payload, lock_key), daemon=True)
                            t.start()
                        except Exception as e:
                            try:
                                cache.delete(lock_key)
                            except Exception:
                                pass
                    else:
                        # 未获取到锁，检查是否已有锁存在（其他进程在刷新），若存在则返回 syncing
                        try:
                            if cache.get(lock_key):
                                syncing_flag = True
                        except Exception:
                            pass
            except Exception:
                pass

            return drf_ok({'key': cache_key, 'sync_time': cache_created_at, 'needs_refresh': needs_refresh, 'syncing': syncing_flag})
        except Exception as e:
            tb = traceback.format_exc()
            return drf_error('lossmakingorders_sync failed', status=500, data={'msg': str(e), 'trace': tb})


    @action(detail=False, methods=["post"], url_path="lossmakingorders_data")
    def lossmaking_orders_data(self, request):
        """根据 cache key 返回缓存的 pick_fields 数据。
        请求 body (JSON): { key: str, page?: int, page_size?: int }
        返回: { list: [...], total: int, sync_time: str|null }
        """
        # 说明：该接口仅从 `OrderProfitCache` 读取已缓存的 `pick_fields` 数据并按页返回，
        # 不会触发上游请求或对数据做额外过滤（前端若需过滤可在客户端完成或调用其它接口）。
        try:
            if getattr(request, 'method', '').upper() == 'GET':
                payload = request.query_params or {}
            else:
                payload = request.data or {}

            cache_key = payload.get('key')
            if not cache_key:
                return drf_error('key is required', status=400)

            page = int(payload.get('page', 1) or 1)
            page_size = int(payload.get('page_size', 50) or 50)

            data_list = []
            sync_time = None
            try:
                if OrderProfitCache is not None:
                    entry = OrderProfitCache.objects.filter(key=cache_key).first()
                    if entry:
                        try:
                            data_list = json.loads(entry.data or '[]')
                        except Exception:
                            data_list = []
                        try:
                            # 返回最近一次写入时间，使用 updated_at
                            sync_time = entry.updated_at.isoformat()
                        except Exception:
                            sync_time = None
            except Exception:
                data_list = []
            # 额外支持基于 listing 负责人（owners）、MSKU（searchValue）和 rule 的服务器端筛选。
            # 解析可选筛选参数（兼容字符串或数组）
            owners = payload.get('owners')
            if isinstance(owners, str):
                try:
                    owners = json.loads(owners)
                except Exception:
                    owners = [x.strip() for x in owners.split(',') if x.strip()]

            search_value = payload.get('searchValue') or payload.get('msku')
            if isinstance(search_value, str):
                # 支持多行输入
                search_value = [s.strip() for s in search_value.splitlines() if s.strip()]

            # 可选的 sids 过滤：前端在 data 阶段可传入 sids 列表用于基于 item.sids 做服务器端过滤
            sids_filter = payload.get('sids')
            if isinstance(sids_filter, str):
                try:
                    sids_filter = json.loads(sids_filter)
                except Exception:
                    try:
                        sids_filter = [int(x) for x in sids_filter.split(',') if x.strip()]
                    except Exception:
                        sids_filter = [x.strip() for x in sids_filter.split(',') if x.strip()]

            rule_name = payload.get('rule') or None

            # 支持排序参数：sort_by 和 sort_order（'asc' / 'desc'）
            sort_by = payload.get('sort_by')
            sort_order = (payload.get('sort_order') or 'asc').lower()

            def match_owners(item: dict, owners_list) -> bool:
                if not owners_list:
                    return True
                principals = []
                # 首选从 price_list 中提取 principal_uids（可能为 list、单值或分隔字符串）
                for p in (item.get('price_list') or []):
                    try:
                        if not isinstance(p, dict):
                            continue
                        u = None
                        if 'principal_uids' in p:
                            u = p.get('principal_uids')
                        elif 'principal_uid' in p:
                            u = p.get('principal_uid')
                        else:
                            # 兼容不同大小写或命名的字段
                            for k in p.keys():
                                if k.lower() in ('principal_uid', 'principaluids', 'principal_uids', 'principaluid'):
                                    u = p.get(k)
                                    break
                        if u is None:
                            continue
                        if isinstance(u, list):
                            for x in u:
                                if x is not None:
                                    principals.append(str(x).strip())
                        else:
                            # 可能为带分隔符的字符串或单一值
                            us = str(u)
                            if re.search(r'[,;\n]', us):
                                principals.extend([s.strip() for s in re.split(r'[,\n;]+', us) if s.strip()])
                            else:
                                if us.strip():
                                    principals.append(us.strip())
                    except Exception:
                        continue

                # 将 owners_list 中的任意项与 principals 精确比较（字符串化后比对）
                for o in owners_list:
                    for p in principals:
                        if str(o).strip() == str(p).strip():
                            return True
                return False

            def match_msku(item: dict, msku_list) -> bool:
                if not msku_list:
                    return True
                # 从 price_list/local_infos 中提取可能的 MSKU 值进行匹配
                candidates = []
                for p in (item.get('price_list') or []):
                    try:
                        if isinstance(p, dict):
                            if p.get('local_sku'):
                                candidates.append(str(p.get('local_sku')))
                            if p.get('sku'):
                                candidates.append(str(p.get('sku')))
                    except Exception:
                        continue
                for li in (item.get('local_infos') or []):
                    try:
                        if isinstance(li, dict) and li.get('local_sku'):
                            candidates.append(str(li.get('local_sku')))
                    except Exception:
                        continue
                # 精确匹配任意一项
                for q in msku_list:
                    for c in candidates:
                        if str(q).strip() == str(c).strip():
                            return True
                return False

            def match_sids(item: dict, sids_list) -> bool:
                if not sids_list:
                    return True
                item_sids = item.get('sids') or []
                vals = []
                try:
                    if isinstance(item_sids, str):
                        vals = [s.strip() for s in re.split(r'[,;\n]+', item_sids) if s.strip()]
                    elif isinstance(item_sids, list):
                        vals = [str(x).strip() for x in item_sids if x is not None]
                    else:
                        vals = [str(item_sids).strip()]
                except Exception:
                    vals = [str(item_sids)]
                for q in sids_list:
                    for v in vals:
                        if str(q).strip() == str(v).strip():
                            return True
                return False

            # 应用 owners / MSKU 过滤
            try:
                filtered = []
                for it in data_list:
                    if not match_owners(it, owners):
                        continue
                    if not match_msku(it, search_value):
                        continue
                    if not match_sids(it, sids_filter):
                        continue
                    filtered.append(it)
            except Exception:
                filtered = data_list

            # 应用规则过滤（由 apply_rule 处理，当前为占位实现）
            try:
                filtered = self.apply_rule(filtered, rule_name)
            except Exception:
                pass

            # 应用排序（服务器端）
            try:
                if sort_by:
                    allowed = {'gross_profit', 'gross_margin', 'net_gross_margin', 'return_rate', 'refund_amount_rate', 'total_stock_fee', 'spend', 'spend_rate'}
                    if str(sort_by) in allowed:
                        # 为了保证 None 值始终排在末尾，按不同排序方向构造 key
                        if str(sort_order) == 'desc':
                            def sort_key(it):
                                v = it.get(sort_by)
                                n = self.parse_number(v)
                                if n is None:
                                    return (1, 0)
                                return (0, -n)
                        else:
                            def sort_key(it):
                                v = it.get(sort_by)
                                n = self.parse_number(v)
                                if n is None:
                                    return (1, 0)
                                return (0, n)
                        try:
                            filtered.sort(key=sort_key)
                        except Exception:
                            pass
            except Exception:
                pass

            total = len(filtered)
            start_idx = max(0, (page - 1) * page_size)
            end_idx = start_idx + page_size
            page_items = filtered[start_idx:end_idx]

            return drf_ok({'list': page_items, 'total': total, 'sync_time': sync_time})
        except Exception as e:
            tb = traceback.format_exc()
            return drf_error('lossmakingorders_data failed', status=500, data={'msg': str(e), 'trace': tb})


class MonthlyLossViewSet(viewsets.ViewSet):
    """Monthly loss orders (CRUD + filter by month/owner, owner optional). Parameters and JSON responses use English keys only."""

    def get_permissions(self):
        method = getattr(self.request, 'method', '').upper() if hasattr(self, 'request') else ''
        if method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]

    @action(detail=False, methods=["get", "post"], url_path="")
    def list_or_create(self, request):
        if request.method.lower() == 'get':
            qs = MonthlyLossOrder.objects.all().order_by('-month', '-id')
            month_q = request.query_params.get('month')
            owner_q = request.query_params.get('owner')
            # 接受 YYYYMM 或 YYYY-MM 格式输入；匹配数据库中可能为 'YYYYMM' 或 'YYYY-MM' 的值
            def _month_variants(m):
                if not m:
                    return []
                s = str(m).strip()
                if re.match(r'^\d{6}$', s):
                    return [s, f"{s[:4]}-{s[4:6]}"]
                if re.match(r'^\d{4}-\d{2}$', s):
                    return [s, s.replace('-', '')]
                return [s]
            try:
                if month_q:
                    variants = _month_variants(month_q)
                    from django.db.models import Q
                    qf = Q()
                    for v in variants:
                        qf |= Q(month__iexact=v)
                    qs = qs.filter(qf)
            except Exception:
                pass
            try:
                if owner_q:
                    qs = qs.filter(owner__iexact=str(owner_q).strip())
            except Exception:
                pass
            total, items, _, _ = paginate_queryset(request, qs)
            data = MonthlyLossSerializer(items, many=True).data
            return drf_ok({'total': total, 'list': data})
        # POST -> 创建
        payload = request.data or {}
        ser = MonthlyLossSerializer(data=payload)
        if not ser.is_valid():
            return drf_error('参数错误', status=400, data={'errors': ser.errors})
        obj = ser.save()
        return drf_ok(MonthlyLossSerializer(obj).data, status=201)

    @action(detail=False, methods=["get"], url_path=r"(?P<id>[^/]+)/form")
    def form(self, request, id: str):
        try:
            obj = MonthlyLossOrder.objects.get(pk=id)
        except MonthlyLossOrder.DoesNotExist:
            return drf_error("未找到记录", status=404)
        return drf_ok(MonthlyLossSerializer(obj).data)

    @action(detail=False, methods=["put", "delete"], url_path=r"(?P<ids>[^/]+)")
    def update_or_delete(self, request, ids: str):
        if request.method.lower() == 'put':
            first_id = ids.split(',')[0]
            try:
                obj = MonthlyLossOrder.objects.get(pk=first_id)
            except MonthlyLossOrder.DoesNotExist:
                return drf_error("未找到记录", status=404)
            ser = MonthlyLossSerializer(obj, data=request.data or {}, partial=True)
            if not ser.is_valid():
                return drf_error('参数错误', status=400, data={'errors': ser.errors})
            obj = ser.save()
            return drf_ok(MonthlyLossSerializer(obj).data)
        # 删除
        id_list = [i for i in ids.split(',') if i]
        MonthlyLossOrder.objects.filter(id__in=id_list).delete()
        return drf_ok(status=204)

    @action(detail=False, methods=["get", "post"], url_path="download")
    def download(self, request):
        # 更简洁的单线程导出：按产品键分块遍历并写入行
        try:
            if getattr(request, 'method', '').upper() == 'GET':
                payload = request.query_params or {}
            else:
                payload = request.data or {}
            owner_q = payload.get('owner')
            time_range = payload.get('time') or payload.get('months')
            if not time_range:
                return drf_error('time parameter required, example: 202610-202612', status=400)

            months = parse_months(str(time_range))
            # 支持区间或单月输入；若 parse_months 返回空，则尝试解析单月格式
            if not months:
                single = norm_month(str(time_range))
                if single:
                    months = [single]
            if not months:
                return drf_error('please provide valid month or month range, example: 202510 or 202510-202512', status=400)

            # 准备缓存 key（保持一致），并尝试从缓存快速返回
            batch_size = int(payload.get('batch_size', 500) or 500)
            cache_key = build_cache_key(owner_q, time_range, parse_store_param(payload.get('store') or payload.get('stores')), months, batch_size)


            # 支持请求参数触发强制刷新缓存（refresh/refresh_cache/force_refresh）
            refresh_flag = is_refresh_requested(payload)
            if cache_key:
                if refresh_flag:
                    remove_cache_and_disk(cache_key)
                try:
                    if not refresh_flag:
                        resp = try_get_cached_file_response(cache_key, months)
                        if resp:
                            return resp
                except Exception:
                    pass

            # 接受 `store`（首选）或兼容旧的 `stores` 参数；
            # 过滤规则：当任意提供的店铺值包含于 `store_country` 字段时匹配
            stores = parse_store_param(payload.get('store') or payload.get('stores'))

            variant_months = set()
            for m in months:
                variant_months.add(m); variant_months.add(m.replace('-', ''))
            qs = MonthlyLossOrder.objects.filter(month__in=list(variant_months))
            if owner_q:
                qs = qs.filter(owner__iexact=str(owner_q).strip())
            if stores:
                from django.db.models import Q
                q = Q()
                for s in stores:
                    if not s:
                        continue
                    q |= Q(store_country__icontains=s)
                qs = qs.filter(q)

            try:
                if qs.count() == 0:
                    return drf_ok(status=204)
            except Exception:
                pass

            from django.db import close_old_connections

            metrics = [('sales','sales'), ('gross_profit','gross_profit'), ('gross_margin','gross_margin'), ('net_gross_margin','net_gross_margin'), ('return_rate','return_rate'), ('refund_rate','refund_amount_rate'), ('storage_fee','total_stock_fee'), ('ad_cost','spend'), ('ad_rate','spend_rate')]

            # 将 product keys 列表化，避免 MySQL 嵌套游标导致的冲突
            product_keys = list(qs.values_list('msku','asin','parent_asin','store_country').distinct())
            def chunked_iter(seq, size=500):
                batch = []
                for d in seq:
                    msku, asin, parent_asin, store_country = d
                    key = (msku or '', asin or '', parent_asin or '', store_country or '')
                    batch.append(key)
                    if len(batch) >= size:
                        yield batch
                        batch = []
                if batch:
                    yield batch


            # 单次聚合流程：一次性批量获取匹配行，在内存中聚合，然后写入 xlsx
            import tempfile, os, threading, shutil

            batch_size = int(payload.get('batch_size', 500) or 500)
            close_old_connections()

            try:
                # 一次性批量查询所有相关行（可能很大；请确保服务器有足够内存）
                rows_q = list(qs.values(
                    'msku','asin','parent_asin','store_country',
                    'product_name_sku','owner','month','sales','gross_profit','gross_margin',
                    'net_gross_margin','return_rate','refund_amount_rate','total_stock_fee','spend','spend_rate'
                ))

                # 使用 pandas 按产品与月份进行聚合（更快、更简洁）
                try:
                    import pandas as pd
                except Exception:
                    return drf_error('pandas is required for optimized aggregation, please install it', status=500)

                try:
                    df = pd.DataFrame(rows_q)
                    if df.empty:
                        aggregated = {}
                    else:
                        # 规范化字段：month -> YYYY-MM，msku/parent_asin/store_country 做清洗
                        df['month_norm'] = df.get('month').apply(norm_month)

                        # 确保数值列为数值类型
                        num_cols = ['sales','gross_profit','gross_margin','net_gross_margin','return_rate','refund_amount_rate','total_stock_fee','spend','spend_rate']
                        for c in num_cols:
                            if c in df.columns:
                                df[c] = pd.to_numeric(df[c], errors='coerce')

                        group_cols = ['msku','asin','parent_asin','store_country','month_norm']
                        agg_dict = {
                            'product_name_sku': lambda x: next((v for v in x if pd.notna(v) and str(v).strip()), ''),
                            'owner': lambda x: next((v for v in x if pd.notna(v) and str(v).strip()), ''),
                            'sales': 'sum',
                            'gross_profit': 'sum',
                            'total_stock_fee': 'sum',
                            'spend': 'sum',
                            'gross_margin': 'mean',
                            'net_gross_margin': 'mean',
                            'return_rate': 'mean',
                            'refund_amount_rate': 'mean',
                            'spend_rate': 'mean',
                        }

                        grp = df.groupby(group_cols).agg(agg_dict).reset_index()

                        # 构建与之前结构类似的聚合字典
                        aggregated = {}
                        for _, rowg in grp.iterrows():
                            try:
                                msku = rowg.get('msku') or ''
                                asin = rowg.get('asin') or ''
                                parent_asin = rowg.get('parent_asin') or ''
                                store_country = rowg.get('store_country') or ''
                                month = rowg.get('month_norm')
                                key = (msku, asin, parent_asin, store_country)
                                entry = aggregated.setdefault(key, {'meta': {'product_name': '', 'owner': ''}, 'months': {m: {} for m in months}})
                                if rowg.get('product_name_sku'):
                                    entry['meta']['product_name'] = rowg.get('product_name_sku')
                                if rowg.get('owner'):
                                    entry['meta']['owner'] = rowg.get('owner')
                                if month is None:
                                    continue
                                # 赋值聚合后的指标
                                entry['months'][month] = {
                                    'sales': int(rowg['sales']) if pd.notna(rowg['sales']) else None,
                                    'gross_profit': float(rowg['gross_profit']) if pd.notna(rowg['gross_profit']) else None,
                                    'gross_margin': float(rowg['gross_margin']) if pd.notna(rowg['gross_margin']) else None,
                                    'net_gross_margin': float(rowg['net_gross_margin']) if pd.notna(rowg['net_gross_margin']) else None,
                                    'return_rate': float(rowg['return_rate']) if pd.notna(rowg['return_rate']) else None,
                                    'refund_amount_rate': float(rowg['refund_amount_rate']) if pd.notna(rowg['refund_amount_rate']) else None,
                                    'total_stock_fee': float(rowg['total_stock_fee']) if pd.notna(rowg['total_stock_fee']) else None,
                                    'spend': float(rowg['spend']) if pd.notna(rowg['spend']) else None,
                                    'spend_rate': float(rowg['spend_rate']) if pd.notna(rowg['spend_rate']) else None,
                                }
                            except Exception:
                                continue
                except Exception:
                    # 如果 pandas 异常，则回退为空的聚合结果
                    aggregated = {}

                # 从聚合后的数据生成最终的 xlsx 文件
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                tmp_name = tmp.name
                tmp.close()

                try:
                    from openpyxl import Workbook
                except Exception:
                    try:
                        os.remove(tmp_name)
                    except Exception:
                        pass
                    return drf_error('openpyxl is required for xlsx export, please install it', status=500)

                # 使用普通 Workbook 以支持合并单元格与样式
                wb = Workbook()
                ws = wb.active
                ws.title = 'monthly_loss'

                # 中文表头映射
                basic_cols = ['MSKU','ASIN','parent_asin','store_country','product_name','owner']
                basic_map = {
                    'MSKU': 'MSKU',
                    'ASIN': 'ASIN',
                    'parent_asin': '父ASIN',
                    'store_country': '店铺',
                    'product_name': '品名/SKU',
                    'owner': '负责人',
                }
                metric_map = {
                    'sales': '销量',
                    'gross_profit': '毛利',
                    'gross_margin': '毛利率',
                    'net_gross_margin': '净毛利率',
                    'return_rate': '退货率',
                    'refund_rate': '退款率',
                    'storage_fee': '仓储费',
                    'ad_cost': '广告花费',
                    'ad_rate': '广告占比',
                }

                # 样式与边框策略：深线用于不同内容/分组分隔，浅线用于月份内列分隔，整体对比色为黑色
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                header_font = Font(bold=True)
                header_fill = PatternFill(fill_type='solid', start_color='FFDCE6F1', end_color='FFDCE6F1')
                center = Alignment(horizontal='center', vertical='center')
                # 定义边框线：深线（分组边界）、浅线（月内边界）、外框（黑色）
                dark_side = Side(border_style='medium', color='FF000000')
                light_side = Side(border_style='thin', color='FFBBBBBB')
                outer_side = Side(border_style='medium', color='FF000000')
                default_side = Side(border_style='thin', color='FF000000')

                # 预创建并缓存 Border 实例以复用，避免大量重复对象导致的比较开销
                side_map = {
                    'dark': dark_side,
                    'light': light_side,
                    'outer': outer_side,
                    'default': default_side,
                }
                _border_cache = {}
                def get_border(l_key='default', r_key='default', t_key='default', b_key='default'):
                    key = (l_key, r_key, t_key, b_key)
                    if key in _border_cache:
                        return _border_cache[key]
                    left = side_map.get(l_key, default_side)
                    right = side_map.get(r_key, default_side)
                    top = side_map.get(t_key, default_side)
                    bottom = side_map.get(b_key, default_side)
                    b = Border(left=left, right=right, top=top, bottom=bottom)
                    _border_cache[key] = b
                    return b

                # 构造两行表头并记录分组边界
                col_idx = 1
                metric_groups = []
                basic_count = len(basic_cols)
                # 先写基本列，并垂直合并两行；在基本列末尾使用深线作为分隔
                for i, c in enumerate(basic_cols):
                    ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
                    cell = ws.cell(row=1, column=col_idx, value=basic_map.get(c, c))
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center
                    # 如果是基本列最后一列，在右侧使用深线分隔；上下使用深线
                    if i == basic_count - 1:
                        cell.border = get_border('outer', 'dark', 'outer', 'outer')
                    else:
                        cell.border = get_border('outer', 'default', 'outer', 'outer')
                    col_idx += 1

                # 接着为每个指标写合并的顶层单元格，并在第二行写各月列，同时记录各组起止列
                for gi, (mname, _) in enumerate(metrics):
                    label_cn = metric_map.get(mname, mname)
                    start_col = col_idx
                    end_col = col_idx + len(months) - 1
                    metric_groups.append((start_col, end_col))
                    # 合并第一行的月份组标签
                    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
                    top_cell = ws.cell(row=1, column=start_col, value=label_cn)
                    top_cell.font = header_font
                    top_cell.fill = header_fill
                    top_cell.alignment = center
                    # 顶层组单元格左右两端使用深线，顶部/底部使用深线
                    top_cell.border = get_border(('dark' if start_col == basic_count + 1 else 'default'), 'dark', 'outer', 'outer')
                    # 第二行写月份小标题，组内使用浅线分隔，组首使用深线左边界，组尾使用深线右边界
                    for i, m in enumerate(months):
                        cidx = start_col + i
                        cell2 = ws.cell(row=2, column=cidx, value=m)
                        cell2.font = header_font
                        cell2.fill = header_fill
                        cell2.alignment = center
                        left_key = 'dark' if i == 0 else 'light'
                        right_key = 'dark' if i == (len(months) - 1) else 'light'
                        cell2.border = get_border(left_key, right_key, 'outer', 'outer')
                    col_idx = end_col + 1

                total_cols = col_idx - 1

                # 写入数据行，从第三行开始；写入后为每个单元格应用边框样式以区分分组与月份
                rows_written = 0
                for key, entry in aggregated.items():
                    try:
                        msku, asin, parent_asin, store_country = key
                        meta = entry.get('meta', {})
                        row = [msku, asin, parent_asin, store_country, meta.get('product_name',''), meta.get('owner','')]
                        # 比率字段格式化为百分比显示
                        for _, field in metrics:
                            for m in months:
                                v = entry['months'].get(m, {}).get(field)
                                if v is None or v == '':
                                    row.append('')
                                    continue
                                if field in RATIO_FIELDS:
                                    row.append(format_ratio_value(v))
                                else:
                                    row.append(v)
                        ws.append(row)
                        rows_written += 1
                        # 应用边框：根据列位置决定左右边框是深线或浅线，顶部/底部使用默认黑色细线
                        # 避免使用 ws.max_row（在某些环境下可能触发 empty _cells 错误），使用已追踪的行号
                        r = 2 + rows_written
                        for c in range(1, total_cols + 1):
                            cell = ws.cell(row=r, column=c)
                            # 初始化左右边框为默认细黑线
                            left = default_side
                            right = default_side
                            # 若在 basic 列内部，basic 与 metrics 交界处使用深线
                            if c == basic_count:
                                border_obj = get_border('default', 'dark', 'default', 'default')
                            else:
                                # 判断是否为任一 metric group 的首列或末列
                                is_group_edge = False
                                left_key = 'default'
                                right_key = 'default'
                                for (gstart, gend) in metric_groups:
                                    if c == gstart:
                                        left_key = 'dark'
                                        is_group_edge = True
                                        break
                                    if c == gend:
                                        right_key = 'dark'
                                        is_group_edge = True
                                        break
                                if not is_group_edge:
                                    # 非边界列使用浅线分隔
                                    left_key = 'light'
                                    right_key = 'light'
                                border_obj = get_border(left_key, right_key, 'default', 'default')
                            cell.border = border_obj
                            # 对月份列（metric 列）按规则着色：对每个 month 做一次判断并为该月所有 metric 单元格上色
                            try:
                                if c > basic_count:
                                    cols_per_metric = len(months)
                                    offset = c - basic_count - 1
                                    month_idx = offset % cols_per_metric
                                    month_key = months[month_idx]
                                    month_vals = entry.get('months', {}).get(month_key, {})
                                    color = determine_month_color(month_vals)
                                    if color:
                                        try:
                                            metric_idx = offset // cols_per_metric
                                            metric_field = metrics[metric_idx][1] if 0 <= metric_idx < len(metrics) else None
                                        except Exception:
                                            metric_field = None
                                        # 不为销量、仓储费、广告花费的列才上色
                                        if metric_field in ('sales', 'total_stock_fee', 'spend'):
                                            pass
                                        else:
                                            cell.fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
                            except Exception:
                                pass
                    except Exception:
                        continue

                if rows_written == 0:
                    try:
                        total_cols = col_idx - 1
                    except Exception:
                        total_cols = len(basic_cols) + len(metrics) * len(months)
                    note = ['No data for requested months'] + [''] * (total_cols - 1)
                    try:
                        ws.append(note)
                    except Exception:
                        pass

                try:
                    wb.save(tmp_name)
                    try:
                        with open(tmp_name, 'rb') as rf:
                            data_bytes = rf.read()
                    except Exception:
                        data_bytes = None
                except Exception:
                    try:
                        os.remove(tmp_name)
                    except Exception:
                        pass
                    raise

                filename = f"monthly_loss_{months[0].replace('-','')}_{months[-1].replace('-','')}.xlsx"

                # 尝试把字节缓存到后端（使用后端默认过期策略），失败时回退到磁盘并缓存 path
                if cache_key and data_bytes is not None:
                    try:
                        cache_data_bytes_with_fallback(cache_key, data_bytes, filename)
                    except Exception:
                        pass

                # 返回临时文件流并在后台清理（封装在 helper 中）
                resp = stream_tempfile_response(tmp_name, filename)
                if resp:
                    return resp
                # 若 helper 失败，回退为原始实现
                from django.http import FileResponse
                f = open(tmp_name, 'rb')
                resp = FileResponse(f, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                resp['Content-Disposition'] = f'attachment; filename="{filename}"'
                def _cleanup_file(path_file, delay=30):
                    try:
                        time.sleep(delay)
                        try:
                            os.remove(path_file)
                        except Exception:
                            pass
                    except Exception:
                        pass
                t = threading.Thread(target=_cleanup_file, args=(tmp_name,))
                t.daemon = True
                t.start()

                return resp
            except Exception as e:
                try:
                    tmp_name
                except Exception:
                    tmp_name = None
                try:
                    if tmp_name:
                        os.remove(tmp_name)
                except Exception:
                    pass
                tb = traceback.format_exc()
                return drf_error('failed to generate xlsx', status=500, data={'msg': str(e), 'trace': tb})
        except Exception as e:
            tb = traceback.format_exc()
            return drf_error('download failed', status=500, data={'msg': str(e), 'trace': tb})


class MonthlyLossFirst20ViewSet(viewsets.ViewSet):
    """Monthly first-20-days loss (CRUD + filter by month/owner, owner optional). Parameters and JSON responses use English keys only."""

    def get_permissions(self):
        method = getattr(self.request, 'method', '').upper() if hasattr(self, 'request') else ''
        if method == 'GET':
            return [AllowAny()]
        return [IsAuthenticated()]

    @action(detail=False, methods=["get", "post"], url_path="")
    def list_or_create(self, request):
        if request.method.lower() == 'get':
            qs = MonthlyLossOrderFirst20.objects.all().order_by('-month', '-id')
            month_q = request.query_params.get('month')
            owner_q = request.query_params.get('owner')
            # 接受 YYYYMM 或 YYYY-MM 格式输入；匹配数据库中可能为 'YYYYMM' 或 'YYYY-MM' 的值
            def _month_variants(m):
                if not m:
                    return []
                s = str(m).strip()
                if re.match(r'^\d{6}$', s):
                    return [s, f"{s[:4]}-{s[4:6]}"]
                if re.match(r'^\d{4}-\d{2}$', s):
                    return [s, s.replace('-', '')]
                return [s]
            try:
                if month_q:
                    variants = _month_variants(month_q)
                    from django.db.models import Q
                    qf = Q()
                    for v in variants:
                        qf |= Q(month__iexact=v)
                    qs = qs.filter(qf)
            except Exception:
                pass
            try:
                if owner_q:
                    qs = qs.filter(owner__iexact=str(owner_q).strip())
            except Exception:
                pass
            total, items, _, _ = paginate_queryset(request, qs)
            data = MonthlyLossFirst20Serializer(items, many=True).data
            return drf_ok({'total': total, 'list': data})
        # POST -> 创建
        payload = request.data or {}
        ser = MonthlyLossFirst20Serializer(data=payload)
        if not ser.is_valid():
            return drf_error('参数错误', status=400, data={'errors': ser.errors})
        obj = ser.save()
        return drf_ok(MonthlyLossFirst20Serializer(obj).data, status=201)

    @action(detail=False, methods=["get"], url_path=r"(?P<id>[^/]+)/form")
    def form(self, request, id: str):
        try:
            obj = MonthlyLossOrderFirst20.objects.get(pk=id)
        except MonthlyLossOrderFirst20.DoesNotExist:
            return drf_error("未找到记录", status=404)
        return drf_ok(MonthlyLossFirst20Serializer(obj).data)

    @action(detail=False, methods=["put", "delete"], url_path=r"(?P<ids>[^/]+)")
    def update_or_delete(self, request, ids: str):
        if request.method.lower() == 'put':
            first_id = ids.split(',')[0]
            try:
                obj = MonthlyLossOrderFirst20.objects.get(pk=first_id)
            except MonthlyLossOrderFirst20.DoesNotExist:
                return drf_error("未找到记录", status=404)
            ser = MonthlyLossFirst20Serializer(obj, data=request.data or {}, partial=True)
            if not ser.is_valid():
                return drf_error('参数错误', status=400, data={'errors': ser.errors})
            obj = ser.save()
            return drf_ok(MonthlyLossFirst20Serializer(obj).data)
        # 删除
        id_list = [i for i in ids.split(',') if i]
        MonthlyLossOrderFirst20.objects.filter(id__in=id_list).delete()
        return drf_ok(status=204)

    @action(detail=False, methods=["post"], url_path="download")
    def download(self, request):
        # 导出对比：本月前20天数据 vs 上个月整月数据
        try:
            if getattr(request, 'method', '').upper() == 'GET':
                payload = request.query_params or {}
            else:
                payload = request.data or {}

            owner_q = payload.get('owner')
            time_range = payload.get('time') or payload.get('months')
            if not time_range:
                return drf_error('time parameter required, example: 202510', status=400)

            months = parse_months(str(time_range))
            # 支持单月输入（YYYYMM 或 YYYY-MM）以及区间输入
            if not months:
                single = norm_month(str(time_range))
                if single:
                    months = [single]
            # 仅支持单月对比
            if not months or len(months) != 1:
                return drf_error('please provide a single month, format: 202510 or 2025-10', status=400)
            cur_month = months[0]
            # 计算上一个月
            try:
                y = int(cur_month[:4]); m = int(cur_month[-2:])
                pm_y = y if m > 1 else y - 1
                pm_m = m - 1 if m > 1 else 12
                prev_month = f"{pm_y:04d}-{pm_m:02d}"
            except Exception:
                return drf_error('failed to compute previous month', status=400)

            # 准备 cache key（以当前月为主）
            batch_size = int(payload.get('batch_size', 500) or 500)
            cache_key = build_cache_key(owner_q, time_range, parse_store_param(payload.get('store') or payload.get('stores')), [cur_month], batch_size)

            refresh_flag = is_refresh_requested(payload)
            if cache_key and refresh_flag:
                remove_cache_and_disk(cache_key)
            if cache_key and not refresh_flag:
                try:
                    resp = try_get_cached_file_response(cache_key, [cur_month])
                    if resp:
                        return resp
                except Exception:
                    pass

            # 过滤条件
            stores = parse_store_param(payload.get('store') or payload.get('stores'))
            def make_variant_set(m):
                return {m, m.replace('-', '')}

            qs_cur = MonthlyLossOrderFirst20.objects.filter(month__in=list(make_variant_set(cur_month)))
            qs_prev = MonthlyLossOrder.objects.filter(month__in=list(make_variant_set(prev_month)))
            if owner_q:
                qs_cur = qs_cur.filter(owner__iexact=str(owner_q).strip())
                qs_prev = qs_prev.filter(owner__iexact=str(owner_q).strip())
            if stores:
                from django.db.models import Q
                qf = Q()
                for s in stores:
                    if not s:
                        continue
                    qf |= Q(store_country__icontains=s)
                qs_cur = qs_cur.filter(qf)
                qs_prev = qs_prev.filter(qf)

            # 查询并聚合（使用 pandas 与之前相同策略）
            try:
                import pandas as pd
            except Exception:
                return drf_error('pandas is required for optimized aggregation, please install it', status=500)

            rows_cur = list(qs_cur.values('msku','asin','parent_asin','store_country','product_name_sku','owner','month','sales','gross_profit','gross_margin','net_gross_margin','return_rate','refund_amount_rate','total_stock_fee','spend','spend_rate'))
            rows_prev = list(qs_prev.values('msku','asin','parent_asin','store_country','product_name_sku','owner','month','sales','gross_profit','gross_margin','net_gross_margin','return_rate','refund_amount_rate','total_stock_fee','spend','spend_rate'))

            def aggregate_rows(rows, months_list):
                try:
                    df = pd.DataFrame(rows)
                    if df.empty:
                        return {}
                    df['month_norm'] = df.get('month').apply(norm_month)
                    num_cols = ['sales','gross_profit','gross_margin','net_gross_margin','return_rate','refund_amount_rate','total_stock_fee','spend','spend_rate']
                    for c in num_cols:
                        if c in df.columns:
                            df[c] = pd.to_numeric(df[c], errors='coerce')
                    group_cols = ['msku','asin','parent_asin','store_country','month_norm']
                    agg_dict = {
                        'product_name_sku': lambda x: next((v for v in x if pd.notna(v) and str(v).strip()), ''),
                        'owner': lambda x: next((v for v in x if pd.notna(v) and str(v).strip()), ''),
                        'sales': 'sum',
                        'gross_profit': 'sum',
                        'total_stock_fee': 'sum',
                        'spend': 'sum',
                        'gross_margin': 'mean',
                        'net_gross_margin': 'mean',
                        'return_rate': 'mean',
                        'refund_amount_rate': 'mean',
                        'spend_rate': 'mean',
                    }
                    grp = df.groupby(group_cols).agg(agg_dict).reset_index()
                    out = {}
                    for _, rowg in grp.iterrows():
                        try:
                            msku = rowg.get('msku') or ''
                            asin = rowg.get('asin') or ''
                            parent_asin = rowg.get('parent_asin') or ''
                            store_country = rowg.get('store_country') or ''
                            month = rowg.get('month_norm')
                            key = (msku, asin, parent_asin, store_country)
                            entry = out.setdefault(key, {'meta': {'product_name': '', 'owner': ''}, 'months': {}})
                            if rowg.get('product_name_sku'):
                                entry['meta']['product_name'] = rowg.get('product_name_sku')
                            if rowg.get('owner'):
                                entry['meta']['owner'] = rowg.get('owner')
                            if month is None:
                                continue
                            entry['months'][month] = {
                                'sales': int(rowg['sales']) if pd.notna(rowg['sales']) else None,
                                'gross_profit': float(rowg['gross_profit']) if pd.notna(rowg['gross_profit']) else None,
                                'gross_margin': float(rowg['gross_margin']) if pd.notna(rowg['gross_margin']) else None,
                                'net_gross_margin': float(rowg['net_gross_margin']) if pd.notna(rowg['net_gross_margin']) else None,
                                'return_rate': float(rowg['return_rate']) if pd.notna(rowg['return_rate']) else None,
                                'refund_amount_rate': float(rowg['refund_amount_rate']) if pd.notna(rowg['refund_amount_rate']) else None,
                                'total_stock_fee': float(rowg['total_stock_fee']) if pd.notna(rowg['total_stock_fee']) else None,
                                'spend': float(rowg['spend']) if pd.notna(rowg['spend']) else None,
                                'spend_rate': float(rowg['spend_rate']) if pd.notna(rowg['spend_rate']) else None,
                            }
                        except Exception:
                            continue
                    return out
                except Exception:
                    return {}

            agg_cur = aggregate_rows(rows_cur, [cur_month])
            agg_prev = aggregate_rows(rows_prev, [prev_month])

            # 生成 xlsx：基本列 + 每个指标 Prev/Cur
            try:
                from openpyxl import Workbook
            except Exception:
                return drf_error('openpyxl is required for xlsx export, please install it', status=500)

            wb = Workbook()
            ws = wb.active
            ws.title = 'monthly_loss_first20_compare'

            basic_cols = ['MSKU','ASIN','parent_asin','store_country','product_name','owner']
            basic_map = {'MSKU':'MSKU','ASIN':'ASIN','parent_asin':'父ASIN','store_country':'店铺','product_name':'品名/SKU','owner':'负责人'}
            metric_map = {
                'sales': '销量',
                'gross_profit': '毛利',
                'gross_margin': '毛利率',
                'net_gross_margin': '净毛利率',
                'return_rate': '退货率',
                'refund_rate': '退款率',
                'storage_fee': '仓储费',
                'ad_cost': '广告花费',
                'ad_rate': '广告占比',
            }
            metrics = [('sales','sales'),('gross_profit','gross_profit'),('gross_margin','gross_margin'),('net_gross_margin','net_gross_margin'),('return_rate','return_rate'),('refund_rate','refund_amount_rate'),('storage_fee','total_stock_fee'),('ad_cost','spend'),('ad_rate','spend_rate')]

            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_font = Font(bold=True)
            header_fill = PatternFill(fill_type='solid', start_color='FFDCE6F1', end_color='FFDCE6F1')
            center = Alignment(horizontal='center', vertical='center')
            default_side = Side(border_style='thin', color='FF000000')
            light_side = Side(border_style='thin', color='FFBBBBBB')
            dark_side = Side(border_style='medium', color='FF000000')
            side_map = {'dark': dark_side, 'light': light_side, 'default': default_side}
            _border_cache = {}
            def get_border(l='default', r='default', t='default', b='default'):
                key = (l,r,t,b)
                if key in _border_cache:
                    return _border_cache[key]
                bobj = Border(left=side_map.get(l, default_side), right=side_map.get(r, default_side), top=side_map.get(t, default_side), bottom=side_map.get(b, default_side))
                _border_cache[key]=bobj
                return bobj

            # 写表头：两行，第一行为指标名（每指标合并两列），第二行为 Prev/Cur 小标题
            col = 1
            basic_count = len(basic_cols)
            for i,c in enumerate(basic_cols):
                ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
                cell = ws.cell(row=1, column=col, value=basic_map.get(c,c))
                cell.font = header_font; cell.fill = header_fill; cell.alignment = center
                cell.border = get_border('outer','outer','outer','outer') if i==basic_count-1 else get_border('default','default','default','default')
                col += 1

            for mname, _ in metrics:
                start = col
                end = col + 2 - 1
                ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
                label_cn = metric_map.get(mname, mname)
                top = ws.cell(row=1, column=start, value=label_cn)
                top.font = header_font; top.fill = header_fill; top.alignment = center
                # second row: Prev, Cur (不再输出 Diff)
                labels = [prev_month, cur_month]
                for i, lab in enumerate(labels):
                    cidx = start + i
                    cell2 = ws.cell(row=2, column=cidx, value=lab)
                    cell2.font = header_font; cell2.fill = header_fill; cell2.alignment = center
                    cell2.border = get_border('default','default','default','default')
                col = end + 1

            total_cols = col -1

            # 合并键集合
            keys = set(list(agg_cur.keys()) + list(agg_prev.keys()))
            rows_written = 0
            for key in keys:
                try:
                    msku, asin, parent_asin, store_country = key
                    meta = (agg_cur.get(key, {}) .get('meta') or agg_prev.get(key, {}).get('meta') or {})
                    row = [msku, asin, parent_asin, store_country, meta.get('product_name',''), meta.get('owner','')]
                    for _, field in metrics:
                        prev_v = agg_prev.get(key, {}).get('months', {}).get(prev_month, {}).get(field)
                        cur_v = agg_cur.get(key, {}).get('months', {}).get(cur_month, {}).get(field)
                        if field in RATIO_FIELDS:
                            pv = format_ratio_value(prev_v) if prev_v is not None else ''
                            cv = format_ratio_value(cur_v) if cur_v is not None else ''
                            row.extend([pv, cv])
                        else:
                            pv = '' if prev_v is None else prev_v
                            cv = '' if cur_v is None else cur_v
                            row.extend([pv, cv])
                    ws.append(row)
                    rows_written += 1
                    r = 2 + rows_written
                    for c in range(1, total_cols + 1):
                        cell = ws.cell(row=r, column=c)
                        # 计算左右边框：basic 列与 metrics 分组交接处使用深线，组内使用浅线
                        try:
                            if c <= basic_count:
                                left_key = 'default'
                                right_key = 'dark' if c == basic_count else 'default'
                            else:
                                # metric 区域，每个 metric 占两列 (Prev, Cur)
                                offset = c - basic_count - 1
                                metric_idx = offset // 2
                                pos_in_metric = offset % 2  # 0:Prev,1:Cur
                                left_key = 'dark' if pos_in_metric == 0 else 'light'
                                right_key = 'dark' if pos_in_metric == 1 else 'light'
                            cell.border = get_border(left_key, right_key, 'default', 'default')
                        except Exception:
                            try:
                                cell.border = get_border('default','default','default','default')
                            except Exception:
                                pass
                        # 对 Prev/Cur 单元格按规则着色（排除部分字段）
                        try:
                            if c > basic_count:
                                offset = c - basic_count - 1
                                metric_idx = offset // 2
                                col_pos = offset % 2  # 0:Prev,1:Cur
                                field_name = metrics[metric_idx][1] if 0 <= metric_idx < len(metrics) else None
                                if field_name and field_name not in ('sales', 'total_stock_fee', 'spend') and col_pos in (0,1):
                                    month_vals = None
                                    try:
                                        if col_pos == 0:
                                            month_vals = agg_prev.get(key, {}).get('months', {}).get(prev_month, {})
                                        else:
                                            month_vals = agg_cur.get(key, {}).get('months', {}).get(cur_month, {})
                                    except Exception:
                                        month_vals = None
                                    if month_vals:
                                        color = determine_month_color(month_vals)
                                        if color:
                                            try:
                                                cell.fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
                                            except Exception:
                                                pass
                        except Exception:
                            pass
                except Exception:
                    continue

            if rows_written == 0:
                note = ['No data for requested month'] + [''] * (total_cols - 1)
                ws.append(note)

            tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            tmp_name = tmp.name; tmp.close()
            try:
                wb.save(tmp_name)
                with open(tmp_name, 'rb') as rf:
                    data_bytes = rf.read()
            except Exception:
                try:
                    os.remove(tmp_name)
                except Exception:
                    pass
                return drf_error('failed to generate xlsx', status=500)

            filename = f"monthly_loss_first20_compare_{cur_month.replace('-','')}.xlsx"
            if cache_key and data_bytes is not None:
                try:
                    cache_data_bytes_with_fallback(cache_key, data_bytes, filename)
                except Exception:
                    pass

            resp = stream_tempfile_response(tmp_name, filename)
            if resp:
                return resp
            from django.http import FileResponse
            f = open(tmp_name, 'rb')
            resp = FileResponse(f, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="{filename}"'
            def _cleanup(path_file, delay=30):
                try:
                    time.sleep(delay)
                    try:
                        os.remove(path_file)
                    except Exception:
                        pass
                except Exception:
                    pass
            t = threading.Thread(target=_cleanup, args=(tmp_name,))
            t.daemon = True; t.start()
            return resp
        except Exception as e:
            tb = traceback.format_exc()
            return drf_error('download failed', status=500, data={'msg': str(e), 'trace': tb})


__all__ = ["StatisticsViewSet", "MonthlyLossViewSet", "MonthlyLossFirst20ViewSet"]

